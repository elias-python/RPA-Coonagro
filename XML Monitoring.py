import win32com.client
import time
import os
import re
import unicodedata
import ctypes
import pandas as pd
import pythoncom
import threading
import sqlite3
import sys
import zipfile
import xml.etree.ElementTree as ET
from datetime import datetime, timedelta
from copy import copy

# Tenta importar openpyxl para a estética visual do Excel
try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.dataframe import dataframe_to_rows

    FORMATACAO_ATIVA = True
except ImportError:
    FORMATACAO_ATIVA = False
    print("[AVISO] Instale 'openpyxl' para ter a formatação profissional no Excel.")

try:
    import pystray
    from PIL import Image, ImageDraw

    TRAY_ATIVO = True
except ImportError:
    TRAY_ATIVO = False
    print("[AVISO] Instale 'pystray' e 'Pillow' para ativar o ícone de bandeja.")


BLOQUEAR_ATUALIZACAO_SE_SAP_ATIVO = True
STATUS_SAP_BLOQUEANTES = (
    "PROCESSANDO",
    "RODANDO",
    "PAUSADO",
    "PARANDO",
    "INICIANDO",
    "AGUARDANDO IMPRESSAO",
)

INDUSTRIALIZACAO_MONITOR_SPECS = (
    (
        "granulacao",
        "INDUSTRIALIZACAO GRANULACAO",
        "Granulação",
        "#648e43",
    ),
    (
        "mistura_sacaria",
        "INDUSTRIALIZACAO MISTURA SACARIA",
        "Mistura sacaria",
        "#d39b2d",
    ),
    (
        "mistura_fosfatado",
        "INDUSTRIALIZACAO MISTURA FOSFATADO",
        "Mistura fosfatado",
        "#4d7a9f",
    ),
    (
        "mistura_nitrogenado",
        "INDUSTRIALIZACAO MISTURA NITROGENADO",
        "Mistura nitrogenado",
        "#8b5a7a",
    ),
)


# --- CONFIGURAÇÕES ---
# Lê config.json da pasta do script (cada máquina tem o seu)
def _diretorio_aplicacao() -> str:
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


def _caminhos_icone_monitor() -> tuple[str, ...]:
    bases = []
    for base in (
        getattr(sys, "_MEIPASS", ""),
        _diretorio_aplicacao(),
        os.path.dirname(os.path.abspath(__file__)),
    ):
        if base and base not in bases and os.path.isdir(base):
            bases.append(base)

    candidatos = []
    for base in bases:
        for relativo in (
            "robo.ico",
            "robo.png",
            os.path.join("assets", "robo.ico"),
            os.path.join("assets", "robo.png"),
        ):
            caminho = os.path.join(base, relativo)
            if os.path.exists(caminho) and caminho not in candidatos:
                candidatos.append(caminho)
    return tuple(candidatos)


def _aplicar_icone_tk(janela) -> bool:
    for caminho in _caminhos_icone_monitor():
        try:
            if os.path.splitext(caminho)[1].lower() == ".ico":
                janela.iconbitmap(caminho)
                return True

            imagem = janela.tk.PhotoImage(file=caminho)
            janela._rpa_icone_photo = imagem
            janela.iconphoto(True, imagem)
            return True
        except Exception:
            continue
    return False


def _carregar_icone_monitor_pillow():
    if not TRAY_ATIVO:
        return None

    for caminho in _caminhos_icone_monitor():
        try:
            return Image.open(caminho).convert("RGBA")
        except Exception:
            continue
    return None


def _auto_detectar_pasta() -> str:
    """Detecta automaticamente a pasta do SharePoint sem precisar editar config.json."""
    base = os.path.join(os.environ.get("USERPROFILE", ""), "The Mosaic Company")
    candidato = os.path.join(base, "Controladoria PGA1 (Arquivos) - RPA - Coonagro")
    if os.path.isdir(candidato):
        return candidato
    if os.path.isdir(base):
        for sub in os.listdir(base):
            if "RPA - Coonagro" in sub and os.path.isdir(os.path.join(base, sub)):
                return os.path.join(base, sub)
    return ""


def _ler_config() -> dict:
    cfg_path = os.path.join(_diretorio_aplicacao(), "config.json")
    if os.path.exists(cfg_path):
        try:
            import json

            with open(cfg_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception as e:
            print(f"[AVISO] Falha ao ler config.json: {e}. Usando caminho padrão.")
    return {}


def _ler_int_config(nome: str, padrao: int) -> int:
    try:
        return max(0, int(_cfg.get(nome, padrao)))
    except Exception:
        return padrao


NOME_XLSM_ATUAL = "Base_Operacional_Toll_Coonagro.xlsm"
NOME_XLSM_LEGADO = "Base_Dados_Coonagro.xlsm"


def _nomes_xlsm_suportados() -> tuple[str, ...]:
    return (NOME_XLSM_ATUAL, NOME_XLSM_LEGADO)


def _resolver_caminho_xlsm(base_dir: str) -> str:
    for nome in _nomes_xlsm_suportados():
        caminho = os.path.join(base_dir, nome)
        if os.path.exists(caminho):
            return caminho
    return os.path.join(base_dir, NOME_XLSM_ATUAL)


def _pasta_runtime_compartilhada() -> str:
    pasta = os.path.join(pasta_trabalho, "system")
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _resolver_arquivo_compartilhado(nome_arquivo: str) -> str:
    caminho_novo = os.path.join(_pasta_runtime_compartilhada(), nome_arquivo)
    caminho_legado = os.path.join(pasta_trabalho, nome_arquivo)

    if os.path.exists(caminho_novo):
        return caminho_novo

    if os.path.exists(caminho_legado):
        try:
            os.replace(caminho_legado, caminho_novo)
            return caminho_novo
        except Exception:
            return caminho_legado

    return caminho_novo


_cfg = _ler_config()
_cfg_pasta = _cfg.get("pasta_trabalho", "")
if "SEU_USUARIO" in _cfg_pasta or not os.path.isdir(_cfg_pasta):
    _cfg_pasta = _auto_detectar_pasta()
RETENCAO_XMLS_DIAS = _ler_int_config("retencao_xmls_dias", 7)
INTERVALO_LIMPEZA_XML_SEGUNDOS = _ler_int_config("intervalo_limpeza_xml_segundos", 3600)
JANELA_HISTORICO_DIAS = 7
pasta_trabalho = (
    _cfg_pasta
    or r"C:\Users\esantan3\The Mosaic Company\Controladoria PGA1 (Arquivos) - RPA - Coonagro"
)
caminho_db = _resolver_arquivo_compartilhado("dados_rpa_coonagro.db")
caminho_excel = _resolver_caminho_xlsm(pasta_trabalho)
# NFs já exportadas e deletadas — Python não as re-adiciona à planilha
caminho_nfs_lancadas = os.path.join(_diretorio_aplicacao(), "nfs_lancadas.txt")
caminho_xmls_pendentes_recarregar = os.path.join(
    _diretorio_aplicacao(), "xmls_pendentes_recarregar.txt"
)
caminho_aviso_primeiro_uso_outlook = os.path.join(
    _diretorio_aplicacao(), "aviso_primeiro_uso_outlook.txt"
)
pasta_xml = os.path.join(
    pasta_trabalho, "XMLs_Recebidos"
)  # pode ser esvaziada sem perda de dados

os.makedirs(pasta_trabalho, exist_ok=True)
os.makedirs(pasta_xml, exist_ok=True)

# ==========================================
# LOG E ESTADO GLOBAL
# ==========================================
_log_path = _resolver_arquivo_compartilhado("rpa_coonagro.log")
_status_global = "Iniciando..."
_nf_count_global = 0
_notas_novas_global = 0  # NFs novas encontradas nesta sessão (para lançamento no SAP)
_icon_ref = None  # referência ao ícone de bandeja para notificações
_painel_monitor_ref = None
_painel_monitor_lock = threading.Lock()
_aguardando_fechamento_xlsm = (
    False  # evita spam de alerta quando arquivo está bloqueado
)
_proximo_alerta_fechamento_xlsm = 0.0  # epoch para novo alerta após adiamento


def _ler_xmls_pendentes_recarregar() -> int:
    try:
        with open(caminho_xmls_pendentes_recarregar, "r", encoding="utf-8") as f:
            return max(0, int((f.read() or "0").strip()))
    except Exception:
        return 0


def _gravar_xmls_pendentes_recarregar(qtd: int):
    qtd = max(0, int(qtd))
    try:
        if qtd == 0:
            if os.path.exists(caminho_xmls_pendentes_recarregar):
                os.remove(caminho_xmls_pendentes_recarregar)
            return
        with open(caminho_xmls_pendentes_recarregar, "w", encoding="utf-8") as f:
            f.write(str(qtd))
    except Exception:
        pass


def _somar_xmls_pendentes_recarregar(qtd: int) -> int:
    total = _ler_xmls_pendentes_recarregar() + max(0, int(qtd))
    _gravar_xmls_pendentes_recarregar(total)
    return total


def _texto_status_monitor() -> str:
    pendentes = _ler_xmls_pendentes_recarregar()
    if pendentes > 0:
        return (
            f"Notas novas para lancamento: {pendentes}. Aguardando atualizacao da Base."
        )
    return f"Aguardando — {_nf_count_global} NFs no banco"


def _mostrar_card_fechamento_ou_adiar_5min() -> bool:
    """
    Exibe card de decisão quando o xlsm está bloqueado.
    Retorna True quando o usuário escolhe adiar 5 minutos.
    """
    # Tenta abrir card customizado com botoes explicitos.
    try:
        import tkinter as tk

        escolha = {"adiar": False}

        root = tk.Tk()
        _aplicar_icone_tk(root)
        root.withdraw()

        win = tk.Toplevel(root)
        _aplicar_icone_tk(win)
        win.title("RPA Coonagro - Atualizacao da Base")
        win.attributes("-topmost", True)
        win.resizable(False, False)

        frame = tk.Frame(win, padx=14, pady=12)
        frame.pack(fill="both", expand=True)

        lbl = tk.Label(
            frame,
            text=(
                "Ha dados novos para atualizar a base XLSM.\n\n"
                "Feche o arquivo agora para concluir a atualizacao\n"
                "ou adie este aviso por 5 minutos."
            ),
            justify="left",
            anchor="w",
        )
        lbl.pack(fill="x", pady=(0, 10))

        btns = tk.Frame(frame)
        btns.pack(fill="x")

        def _fechar_agora():
            escolha["adiar"] = False
            win.destroy()

        def _adiar_5min():
            escolha["adiar"] = True
            win.destroy()

        b1 = tk.Button(btns, text="Fechar agora", width=16, command=_fechar_agora)
        b1.pack(side="left")

        b2 = tk.Button(btns, text="Adiar 5 min", width=16, command=_adiar_5min)
        b2.pack(side="right")

        def _on_close():
            # Fechar no X conta como adiar para evitar bloqueio de UX.
            escolha["adiar"] = True
            win.destroy()

        win.protocol("WM_DELETE_WINDOW", _on_close)

        win.update_idletasks()
        w = win.winfo_width()
        h = win.winfo_height()
        x = (win.winfo_screenwidth() // 2) - (w // 2)
        y = (win.winfo_screenheight() // 2) - (h // 2)
        win.geometry(f"+{x}+{y}")

        win.grab_set()
        b1.focus_set()
        root.wait_window(win)
        root.destroy()

        return escolha["adiar"]
    except Exception:
        # Fallback: caixa padrao Sim/Nao.
        msg = (
            "Ha dados novos para atualizar a base XLSM.\n\n"
            "Clique em 'Sim' para fechar o arquivo agora e atualizar.\n"
            "Clique em 'Nao' para adiar este aviso por 5 minutos."
        )
        title = "RPA Coonagro - Atualizacao da Base"
        flags = (
            0x00000004  # MB_YESNO
            | 0x00000030  # MB_ICONWARNING
            | 0x00040000  # MB_TOPMOST
            | 0x00010000  # MB_SETFOREGROUND
        )
        resp = ctypes.windll.user32.MessageBoxW(0, msg, title, flags)
        return resp == 7  # IDNO = adiar


def _tentar_fechar_base_xlsm() -> bool:
    """Tenta fechar automaticamente apenas a base XLSM alvo."""

    def _norm(path: str) -> str:
        return os.path.normcase(os.path.normpath(os.path.abspath(path)))

    def _norm_loose(path: str) -> str:
        return str(path or "").replace("/", "\\").strip().lower()

    try:
        pythoncom.CoInitialize()
        excel = win32com.client.GetObject(Class="Excel.Application")
        alvo = _norm(caminho_excel)
        alvo_nome = os.path.basename(alvo)
        alvo_nome_loose = alvo_nome.lower()

        fechado = False

        for wb in excel.Workbooks:
            try:
                nome_wb = os.path.basename(_norm(str(getattr(wb, "Name", ""))))
                full_wb_raw = str(getattr(wb, "FullName", ""))
                caminho_wb = _norm(full_wb_raw) if full_wb_raw else ""
                caminho_wb_loose = _norm_loose(full_wb_raw)

                # Assinaturas permitidas para o caminho da nossa base no ambiente.
                nomes_suportados = tuple(
                    nome.lower() for nome in _nomes_xlsm_suportados()
                )
                assinatura_ok = "rpa - coonagro" in caminho_wb_loose and any(
                    nome in caminho_wb_loose for nome in nomes_suportados
                )

                # Regra de segurança: fecha somente quando o caminho completo casar.
                if caminho_wb and caminho_wb == alvo:
                    try:
                        wb.Close(SaveChanges=True)
                    except Exception:
                        wb.Close(SaveChanges=False)
                    _log(f"[EXCEL] Fechamento seletivo OK: {caminho_wb}")
                    fechado = True
                    break

                # Regra de contingencia (OneDrive/SharePoint):
                # Mesmo nome + assinaturas de pasta/projeto no FullName.
                if str(nome_wb).lower() == alvo_nome_loose and assinatura_ok:
                    try:
                        wb.Close(SaveChanges=True)
                    except Exception:
                        wb.Close(SaveChanges=False)
                    _log(
                        f"[EXCEL] Fechamento seletivo (match flexivel) OK: {full_wb_raw}"
                    )
                    fechado = True
                    break

                # Mesmo nome em outro caminho: explicitamente ignorado.
                if nome_wb == alvo_nome and caminho_wb != alvo:
                    _log(f"[EXCEL] Ignorado (mesmo nome, outro caminho): {caminho_wb}")
            except Exception:
                continue

        if not fechado:
            _log("[EXCEL] Arquivo alvo não encontrado aberto nesta instância do Excel.")
        return fechado
    except Exception as e:
        _log(f"[AVISO] Falha ao tentar fechar xlsm automaticamente: {e}")
    return False


def _resumo_pendencias_tray() -> dict:
    try:
        return _carregar_resumo_monitor(limit_historico=50)
    except Exception:
        return {
            "grupos_pendentes": _notas_novas_global,
            "ultima_extracao_pendente": "",
        }


def _notificar(titulo: str, msg: str):
    """Exibe notificação na bandeja do Windows (requer pystray)."""
    global _icon_ref
    if _icon_ref is not None:
        try:
            _icon_ref.notify(msg, titulo)
        except Exception:
            pass


def _texto_pendentes_tray():
    resumo = _resumo_pendencias_tray()
    pend = int(resumo.get("grupos_pendentes") or 0)
    if pend == 1:
        return "Pendentes para lancamento: 1 nota"
    return f"Pendentes para lancamento: {pend} notas"


def _texto_extracao_pendente_tray():
    resumo = _resumo_pendencias_tray()
    data_extracao = _formatar_datahora_br(resumo.get("ultima_extracao_pendente") or "")
    if data_extracao == "-":
        return "Ultima extracao pendente: -"
    return f"Ultima extracao pendente: {data_extracao}"


def _log(msg):
    ts = time.strftime("%Y-%m-%d %H:%M:%S")
    linha = f"[{ts}] {msg}"
    print(linha)
    try:
        with open(_log_path, "a", encoding="utf-8") as f:
            f.write(linha + "\n")
    except Exception:
        pass


def _mostrar_aviso_primeiro_uso_outlook():
    if os.path.exists(caminho_aviso_primeiro_uso_outlook):
        return

    msg = (
        "Primeiro acesso do Extrator XML nesta maquina.\n\n"
        "Antes de usar, crie no Outlook uma pasta com o nome exato 'XML Coonagro'.\n"
        "E dessa pasta que o sistema vai ler e extrair os arquivos XML recebidos.\n\n"
        "Depois de criar a pasta, mova para ela os e-mails com XML que devem entrar no fluxo."
    )
    title = "RPA Coonagro - Configuracao inicial do Outlook"
    flags = (
        0x00000040  # MB_ICONINFORMATION
        | 0x00000000  # MB_OK
        | 0x00040000  # MB_TOPMOST
        | 0x00010000  # MB_SETFOREGROUND
    )

    try:
        ctypes.windll.user32.MessageBoxW(0, msg, title, flags)
    except Exception:
        _log("[AVISO] Nao foi possivel exibir o aviso inicial da pasta Outlook.")
        return

    try:
        with open(caminho_aviso_primeiro_uso_outlook, "w", encoding="utf-8") as f:
            f.write(time.strftime("%Y-%m-%d %H:%M:%S"))
    except Exception as e:
        _log(f"[AVISO] Falha ao gravar marcador do aviso inicial do Outlook: {e}")


# ==========================================
# GESTÃO DO BANCO DE DADOS (SQLITE)
# ==========================================


def inicializar_db():
    conn = sqlite3.connect(caminho_db)
    cursor = conn.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS notas_itens (
            chave_acesso TEXT,
            nf TEXT,
            seq INTEGER,
            cod_material TEXT,
            descricao TEXT,
            ordem_producao TEXT,
            qtd REAL,
            un TEXT,
            vlr_unit REAL,
            vlr_total_item REAL,
            cfop TEXT,
            nf_vinculada_5124 TEXT,
            nf_origem TEXT,
            serie_origem TEXT,
            vlr_total_nf REAL,
            emissao TEXT,
            data_importacao TEXT,
            PRIMARY KEY (chave_acesso, seq)
        )
    """)
    # Migra banco existente (adiciona coluna se ainda nao existir)
    try:
        conn.execute("ALTER TABLE notas_itens ADD COLUMN data_importacao TEXT")
        conn.commit()
    except Exception:
        pass  # coluna ja existe
    conn.commit()
    conn.close()


# ==========================================
# EXTRAÇÃO E FORMATAÇÃO (ESTILO MOSAIC)
# ==========================================


def safe_float(element):
    """Proteção contra tags vazias ou ausentes no XML"""
    try:
        return float(element.text) if element is not None and element.text else 0.0
    except ValueError:
        return 0.0


def extrair_detalhes_xml(caminho_arquivo):
    itens = []
    try:
        tree = ET.parse(caminho_arquivo)
        root = tree.getroot()

        # Limpeza de Namespace
        for elem in root.iter():
            if "}" in elem.tag:
                elem.tag = elem.tag.split("}", 1)[1]

        n_nf = root.find(".//nNF").text if root.find(".//nNF") is not None else "N/A"
        chave = (
            root.find(".//chNFe").text if root.find(".//chNFe") is not None else "N/A"
        )

        # Data de emissão
        data_emi_tag = root.find(".//dhEmi")
        data_emissao = (
            data_emi_tag.text[:10]
            if data_emi_tag is not None and data_emi_tag.text
            else "N/A"
        )

        v_total_nf = safe_float(root.find(".//vNF"))

        # Extração de Referências (Mãe e Sequenciais)
        ref_mae = "N/A"
        refs_origem = []
        inf_cpl = root.find(".//infCpl")
        if inf_cpl is not None and inf_cpl.text:
            m_mae = re.search(r"Ref\. NF-e (\d+)", inf_cpl.text, re.IGNORECASE)
            if m_mae:
                ref_mae = str(int(m_mae.group(1)))
            m_orig = re.findall(
                r"Nota de referência\s+(\d+)\s*-\s*(\d+)", inf_cpl.text, re.IGNORECASE
            )
            if m_orig:
                refs_origem = [(str(int(n)), s) for n, s in m_orig]

        # Loop pelos itens da nota
        for idx, item in enumerate(root.findall(".//det"), start=1):
            prod = item.find("prod")
            if prod is not None:
                cfop = (
                    prod.find("CFOP").text if prod.find("CFOP") is not None else "N/A"
                )

                # Filtra: somente CFOP 5902 (materiais) e 5124 (NF mãe)
                if cfop not in ("5902", "5124"):
                    continue
                ref_item = (
                    refs_origem[idx - 1][0]
                    if (cfop == "5902" and idx - 1 < len(refs_origem))
                    else "N/A"
                )
                ser_item = (
                    str(
                        int(refs_origem[idx - 1][1])
                    )  # remove zeros a esquerda (ex: "013" -> "13")
                    if (
                        cfop == "5902"
                        and idx - 1 < len(refs_origem)
                        and refs_origem[idx - 1][1].isdigit()
                    )
                    else (
                        refs_origem[idx - 1][1]
                        if (cfop == "5902" and idx - 1 < len(refs_origem))
                        else "N/A"
                    )
                )

                itens.append(
                    (
                        chave,
                        n_nf,
                        idx,
                        (
                            prod.find("cProd").text
                            if prod.find("cProd") is not None
                            else "N/A"
                        ),
                        (
                            prod.find("xProd").text
                            if prod.find("xProd") is not None
                            else "N/A"
                        ),
                        (
                            prod.find("xPed").text
                            if prod.find("xPed") is not None
                            else "N/A"
                        ),
                        safe_float(prod.find("qCom")),
                        (
                            prod.find("uCom").text
                            if prod.find("uCom") is not None
                            else "N/A"
                        ),
                        safe_float(prod.find("vUnCom")),
                        safe_float(prod.find("vProd")),
                        cfop,
                        ref_mae if cfop == "5902" else "N/A",
                        ref_item,
                        ser_item,
                        v_total_nf,
                        data_emissao,
                    )
                )
        return itens
    except Exception as e:
        _log(f"[ERRO EXTRAÇÃO] Falha no XML {os.path.basename(caminho_arquivo)}: {e}")
        return None


def _pode_limpar_xml_antigo(caminho_arquivo, chaves_processadas) -> bool:
    if RETENCAO_XMLS_DIAS <= 0:
        return False

    try:
        idade_segundos = time.time() - os.path.getmtime(caminho_arquivo)
    except OSError:
        return False

    if idade_segundos < RETENCAO_XMLS_DIAS * 86400:
        return False

    dados = extrair_detalhes_xml(caminho_arquivo)
    if dados is None:
        return False
    if not dados:
        return True

    try:
        chave = str(dados[0][0]).strip()
    except Exception:
        return False

    return len(chave) > 0 and chave in chaves_processadas


def _limpar_xmls_recebidos_antigos(chaves_processadas) -> int:
    if RETENCAO_XMLS_DIAS <= 0:
        return 0

    removidos = 0
    for arq in os.listdir(pasta_xml):
        if not arq.lower().endswith(".xml"):
            continue

        caminho_arq = os.path.join(pasta_xml, arq)
        if not os.path.isfile(caminho_arq):
            continue

        if not _pode_limpar_xml_antigo(caminho_arq, chaves_processadas):
            continue

        try:
            os.remove(caminho_arq)
            removidos += 1
        except OSError as e:
            _log(f"[LIMPEZA_XML] Falha ao remover {arq}: {e}")

    if removidos > 0:
        _log(
            f"[LIMPEZA_XML] {removidos} XML(s) removido(s) automaticamente "
            f"apos {RETENCAO_XMLS_DIAS} dia(s) de retencao."
        )

    return removidos


def _tentar_substituir_excel(tmp_path, caminho_destino):
    """Move tmp_path para caminho_destino com até 3 tentativas (5s entre elas).
    Se ainda bloqueado, descarta o temp e loga aviso — será tentado no próximo ciclo."""
    import shutil

    global _aguardando_fechamento_xlsm, _status_global, _proximo_alerta_fechamento_xlsm

    if not os.path.exists(tmp_path):
        _log(f"[ERRO] Arquivo temporário não encontrado: {tmp_path}")
        return False

    if not zipfile.is_zipfile(tmp_path):
        _log(
            f"[ERRO] Arquivo temporário inválido (não é um XLSX/XLSM válido): {tmp_path}"
        )
        try:
            os.remove(tmp_path)
        except Exception:
            pass
        return False

    if os.path.exists(caminho_destino):
        destino_tem_vba = _tem_projeto_vba(caminho_destino)
        tmp_tem_vba = _tem_projeto_vba(tmp_path)
        if destino_tem_vba and not tmp_tem_vba:
            _log(
                "[ERRO] Temp sem VBA enquanto destino possui macros. Atualização cancelada para evitar perda de macros."
            )
            try:
                os.remove(tmp_path)
            except Exception:
                pass
            return False

    for tentativa in range(3):
        try:
            if os.path.exists(caminho_destino):
                base_dir = os.path.dirname(caminho_destino)
                nome_base = os.path.splitext(os.path.basename(caminho_destino))[0]
                stamp = time.strftime("%Y%m%d_%H%M%S")
                caminho_backup = os.path.join(
                    base_dir, f"{nome_base}_backup_{stamp}.xlsm"
                )
                shutil.copy2(caminho_destino, caminho_backup)
                _log(f"[EXCEL] Backup criado: {caminho_backup}")
                os.remove(caminho_destino)
            shutil.move(tmp_path, caminho_destino)
            _log("[EXCEL] Relatório atualizado com sucesso.")
            # Se antes estava bloqueado e o usuário fechou, reabre já atualizado.
            if _aguardando_fechamento_xlsm:
                _notificar(
                    "RPA Coonagro — Base Atualizada",
                    "Base atualizada com sucesso. Reabrindo o arquivo atualizado...",
                )
                try:
                    os.startfile(caminho_destino)
                    _log("[EXCEL] Arquivo reaberto automaticamente após atualização.")
                except Exception as e:
                    _log(
                        f"[AVISO] Não foi possível reabrir o xlsm automaticamente: {e}"
                    )
                _aguardando_fechamento_xlsm = False
                _proximo_alerta_fechamento_xlsm = 0.0
                _status_global = "Base atualizada"
            return True
        except PermissionError:
            agora = time.time()
            precisa_alertar = (not _aguardando_fechamento_xlsm) or (
                _proximo_alerta_fechamento_xlsm > 0
                and agora >= _proximo_alerta_fechamento_xlsm
            )

            if precisa_alertar and agora >= _proximo_alerta_fechamento_xlsm:
                _aguardando_fechamento_xlsm = True
                _status_global = "Feche o xlsm para atualizar a base"
                _notificar(
                    "RPA Coonagro — Feche a Base XLSM",
                    f"Novos dados chegaram. Feche '{os.path.basename(caminho_excel)}' para atualizar a base.",
                )

                if _mostrar_card_fechamento_ou_adiar_5min():
                    _proximo_alerta_fechamento_xlsm = agora + 300
                    _status_global = "Atualizacao adiada por 5 minutos"
                    _log("[AVISO] Atualizacao adiada pelo usuario por 5 minutos.")
                else:
                    _proximo_alerta_fechamento_xlsm = 0.0
                    if _tentar_fechar_base_xlsm():
                        _status_global = "Fechando base para atualizar"
                    else:
                        _notificar(
                            "RPA Coonagro — Fechamento Manual Necessario",
                            "Nao consegui fechar a base automaticamente. Feche o arquivo manualmente para atualizar.",
                        )
                    _log(
                        "[AVISO] Usuario optou por fechar o xlsm para atualizar agora."
                    )

            if tentativa < 2:
                _log(
                    f"[AVISO] Excel bloqueado. Nova tentativa em 5s ({tentativa + 2}/3)..."
                )
                time.sleep(5)

    # Esgotou as tentativas — descarta o temp e avisa (próximo ciclo tenta de novo)
    _log("[AVISO] Excel em uso. Feche o arquivo para receber a atualização.")
    try:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    except Exception:
        pass
    return False


def _detectar_linha_cabecalho(ws, max_linhas=5):
    """Detecta em qual linha está o cabeçalho da Base. Retorna None se não encontrar."""
    limite = min(max_linhas, ws.max_row)
    for r in range(1, limite + 1):
        vals = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if "Nº Nota Fiscal" in vals and "Seq." in vals:
            return r
    return None


def _mapa_cabecalho(ws, header_row):
    return {
        ws.cell(header_row, c).value: c
        for c in range(1, ws.max_column + 1)
        if ws.cell(header_row, c).value is not None
    }


def _tem_projeto_vba(path_arquivo):
    """Retorna True se o pacote Office contém projeto VBA."""
    try:
        with zipfile.ZipFile(path_arquivo, "r") as zf:
            return "xl/vbaProject.bin" in zf.namelist()
    except Exception:
        return False


def _ler_nfs_lancadas() -> set:
    """Retorna conjunto de NF numbers já exportadas e deletadas da planilha."""
    nfs: set = set()
    if os.path.exists(caminho_nfs_lancadas):
        try:
            with open(caminho_nfs_lancadas, "r", encoding="utf-8") as f:
                for linha in f:
                    nf = linha.strip()
                    if nf:
                        nfs.add(nf)
        except Exception:
            pass
    return nfs


def _normalizar_nf_texto(valor) -> str:
    txt = str(valor or "").strip()
    if not txt or txt.upper() == "N/A":
        return ""
    try:
        return str(int(float(txt)))
    except Exception:
        return txt


def _normalizar_texto_livre(valor) -> str:
    texto = unicodedata.normalize("NFKD", str(valor or ""))
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", texto.upper()).strip()


def _identificar_tipo_industrializacao(descricao) -> str | None:
    texto = _normalizar_texto_livre(descricao)
    for chave, marcador, _, _ in INDUSTRIALIZACAO_MONITOR_SPECS:
        if marcador in texto:
            return chave
    return None


def _ler_nfs_arquivo(caminho: str) -> set:
    nfs: set = set()
    if not os.path.exists(caminho):
        return nfs
    try:
        with open(caminho, "r", encoding="utf-8") as f:
            for linha in f:
                nf = _normalizar_nf_texto(linha)
                if nf:
                    nfs.add(nf)
    except Exception:
        pass
    return nfs


def _caminho_status_dia() -> str:
    return os.path.join(
        _diretorio_aplicacao(), f"status_dia_{time.strftime('%Y%m%d')}.txt"
    )


def _db_tem_coluna(conn: sqlite3.Connection, tabela: str, coluna: str) -> bool:
    try:
        info = conn.execute(f"PRAGMA table_info({tabela})").fetchall()
        return any(str(item[1]).strip().lower() == coluna.lower() for item in info)
    except Exception:
        return False


def _formatar_toneladas(valor) -> str:
    try:
        numero = float(valor or 0)
    except Exception:
        numero = 0.0
    return f"{numero:,.2f}".replace(",", "_").replace(".", ",").replace("_", ".")


def _formatar_data_br(valor: str) -> str:
    txt = str(valor or "").strip()
    if len(txt) >= 10 and txt[4] == "-" and txt[7] == "-":
        return f"{txt[8:10]}/{txt[5:7]}/{txt[0:4]}"
    return txt or "-"


def _formatar_datahora_br(valor: str) -> str:
    txt = str(valor or "").strip()
    if len(txt) >= 19 and txt[4] == "-" and txt[7] == "-":
        return f"{txt[8:10]}/{txt[5:7]}/{txt[0:4]} {txt[11:16]}"
    if len(txt) >= 10 and txt[4] == "-" and txt[7] == "-":
        return _formatar_data_br(txt)
    return txt or "-"


def _parse_data_monitor(valor: str) -> datetime | None:
    txt = str(valor or "").strip()
    if not txt:
        return None

    candidatos = []
    for candidato in (txt, txt[:19], txt[:16], txt[:10]):
        candidato = candidato.strip()
        if candidato and candidato not in candidatos:
            candidatos.append(candidato)

    formatos = (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%d/%m/%Y %H:%M",
        "%d/%m/%Y",
    )

    for candidato in candidatos:
        for formato in formatos:
            try:
                return datetime.strptime(candidato, formato)
            except ValueError:
                continue
    return None


def _esta_nos_ultimos_dias(valor: str, dias: int) -> bool:
    referencia = _parse_data_monitor(valor)
    if referencia is None:
        return True

    limite = (datetime.now() - timedelta(days=max(0, dias - 1))).date()
    return referencia.date() >= limite


def _abrir_base_xlsm():
    if os.path.exists(caminho_excel):
        os.startfile(caminho_excel)
        return True
    return False


def _abrir_log_monitor():
    if os.path.exists(_log_path):
        os.startfile(_log_path)
        return True
    return False


def _carregar_resumo_monitor(limit_historico: int = 250) -> dict:
    resumo = {
        "status_monitor": _status_global,
        "novas_notas": _ler_xmls_pendentes_recarregar(),
        "grupos_pendentes": 0,
        "toneladas_pendentes": 0.0,
        "grupos_lancados_hoje": 0,
        "toneladas_lancadas_hoje": 0.0,
        "grupos_extraidos_hoje": None,
        "total_grupos_extraidos": 0,
        "industrializacoes_lancadas_hoje": {
            chave: 0 for chave, _, _, _ in INDUSTRIALIZACAO_MONITOR_SPECS
        },
        "grupos_lancados_hoje_sem_tipo": 0,
        "ultima_extracao_pendente": "",
        "historico": [],
    }

    if not os.path.exists(caminho_db):
        return resumo

    nfs_lancadas = {_normalizar_nf_texto(v) for v in _ler_nfs_lancadas()}
    nfs_lancadas.discard("")
    nfs_lancadas_hoje = _ler_nfs_arquivo(_caminho_status_dia())

    conn = sqlite3.connect(caminho_db)
    conn.row_factory = sqlite3.Row
    try:
        tem_data_importacao = _db_tem_coluna(conn, "notas_itens", "data_importacao")
        select_import = (
            ", MAX(data_importacao) AS data_importacao" if tem_data_importacao else ""
        )
        order_import = "MAX(data_importacao) DESC, " if tem_data_importacao else ""
        query = f"""
            SELECT
                nf,
                MAX(emissao) AS emissao,
                MAX(qtd) AS toneladas
                {select_import}
            FROM notas_itens
            WHERE cfop = '5124'
            GROUP BY nf
            ORDER BY {order_import} MAX(emissao) DESC, CAST(nf AS INTEGER) DESC
        """
        rows = conn.execute(query).fetchall()
        resumo["total_grupos_extraidos"] = len(rows)

        detalhes_por_nf = {}
        tipos_por_nf = {}

        tipos_rows = conn.execute("""
            SELECT
                nf,
                descricao
            FROM notas_itens
            WHERE cfop = '5124'
            ORDER BY nf DESC, seq ASC
            """).fetchall()

        for tipo_row in tipos_rows:
            nf_tipo = _normalizar_nf_texto(tipo_row["nf"])
            if not nf_tipo or nf_tipo in tipos_por_nf:
                continue

            tipo_industrializacao = _identificar_tipo_industrializacao(
                tipo_row["descricao"]
            )
            if tipo_industrializacao:
                tipos_por_nf[nf_tipo] = tipo_industrializacao

        detalhes_rows = conn.execute("""
            SELECT
                nf_vinculada_5124,
                nf,
                cod_material,
                descricao,
                qtd,
                un
            FROM notas_itens
            WHERE cfop = '5902'
              AND COALESCE(TRIM(nf_vinculada_5124), '') <> ''
            ORDER BY nf_vinculada_5124 DESC, nf DESC, seq ASC
            """).fetchall()

        for detalhe in detalhes_rows:
            nf_principal = _normalizar_nf_texto(detalhe["nf_vinculada_5124"])
            if not nf_principal:
                continue

            detalhes_por_nf.setdefault(nf_principal, []).append(
                {
                    "nf_par": _normalizar_nf_texto(detalhe["nf"]),
                    "cod_material": str(detalhe["cod_material"] or "").strip(),
                    "descricao": str(detalhe["descricao"] or "").strip(),
                    "toneladas": float(detalhe["qtd"] or 0),
                    "un": str(detalhe["un"] or "").strip(),
                }
            )

        if tem_data_importacao:
            resumo["grupos_extraidos_hoje"] = conn.execute("""
                SELECT COUNT(*)
                FROM (
                    SELECT DISTINCT nf
                    FROM notas_itens
                    WHERE cfop = '5124'
                      AND date(data_importacao) = date('now', 'localtime')
                )
                """).fetchone()[0]

        historico = []
        grupos_pendentes = 0
        grupos_lancados_hoje = 0
        toneladas_pendentes = 0.0
        toneladas_lancadas_hoje = 0.0
        industrializacoes_lancadas_hoje = {
            chave: 0 for chave, _, _, _ in INDUSTRIALIZACAO_MONITOR_SPECS
        }
        grupos_lancados_hoje_sem_tipo = 0

        for row in rows:
            nf = _normalizar_nf_texto(row["nf"])
            if not nf:
                continue

            toneladas = float(row["toneladas"] or 0)
            emissao = str(row["emissao"] or "")
            data_importacao = (
                str(row["data_importacao"] or "") if tem_data_importacao else ""
            )

            if nf in nfs_lancadas_hoje:
                status = "Lançada hoje"
                grupos_lancados_hoje += 1
                toneladas_lancadas_hoje += toneladas
                tipo_industrializacao = tipos_por_nf.get(nf)
                if tipo_industrializacao:
                    industrializacoes_lancadas_hoje[tipo_industrializacao] += 1
                else:
                    grupos_lancados_hoje_sem_tipo += 1
            elif nf in nfs_lancadas:
                status = "Lançada"
            else:
                status = "Pendente"
                grupos_pendentes += 1
                toneladas_pendentes += toneladas
                if not resumo["ultima_extracao_pendente"] and data_importacao:
                    resumo["ultima_extracao_pendente"] = data_importacao

            detalhes_nf = detalhes_por_nf.get(nf, [])
            data_referencia_historico = data_importacao or emissao
            if not _esta_nos_ultimos_dias(
                data_referencia_historico, JANELA_HISTORICO_DIAS
            ):
                continue

            nfs_par = []
            for detalhe in detalhes_nf:
                nf_par = detalhe.get("nf_par") or ""
                if nf_par and nf_par not in nfs_par:
                    nfs_par.append(nf_par)

            historico.append(
                {
                    "nf": nf,
                    "nf_par": " / ".join(nfs_par) if nfs_par else "-",
                    "emissao": emissao,
                    "toneladas": toneladas,
                    "status": status,
                    "data_importacao": data_importacao,
                    "detalhes": detalhes_nf,
                }
            )

        resumo["grupos_pendentes"] = grupos_pendentes
        resumo["toneladas_pendentes"] = toneladas_pendentes
        resumo["grupos_lancados_hoje"] = grupos_lancados_hoje
        resumo["toneladas_lancadas_hoje"] = toneladas_lancadas_hoje
        resumo["industrializacoes_lancadas_hoje"] = industrializacoes_lancadas_hoje
        resumo["grupos_lancados_hoje_sem_tipo"] = grupos_lancados_hoje_sem_tipo
        resumo["historico"] = historico[:limit_historico]
        return resumo
    finally:
        conn.close()


class PainelMonitorXML:
    REFRESH_MS = 5000
    CARD_SPECS = (
        ("grupos_pendentes", "Grupos pendentes", "Fila", "#d39b2d"),
        ("toneladas_pendentes", "Toneladas pendentes", "Ton", "#785b1e"),
        ("grupos_lancados_hoje", "Lançadas hoje", "Dia", "#3f7e58"),
        ("toneladas_lancadas_hoje", "Toneladas hoje", "Peso", "#4d7a9f"),
        ("grupos_extraidos_hoje", "Extraídas hoje", "NF", "#8b5a7a"),
    )

    def __init__(self):
        self.root = None
        self._fechado = False
        self._after_id = None
        self._labels_cards = {}
        self._summary_labels = {}
        self._label_status = None
        self._label_updated = None
        self._label_history_meta = None
        self._history_rows = {}
        self._industrializacao_labels = {}
        self._label_industrializacao_meta = None
        self._details_panel = None
        self._details_title = None
        self._details_meta = None
        self._details_current_iid = None
        self._details_cards_canvas = None
        self._details_cards_frame = None
        self._details_cards_window = None
        self._logo_photo = None
        self._status_dot = None
        self._status_dot_item = None
        self._tree = None

    def aberto(self) -> bool:
        return self.root is not None and not self._fechado

    def executar(self):
        import tkinter as tk
        from tkinter import messagebox, ttk

        self.tk = tk
        self.messagebox = messagebox
        self.ttk = ttk

        root = tk.Tk()
        self.root = root
        _aplicar_icone_tk(root)
        root.title("RPA Coonagro | Central Operacional Toll")
        root.geometry("1380x880")
        root.minsize(1120, 720)
        root.configure(bg="#f4f1e6")
        root.grid_columnconfigure(0, weight=1)
        root.grid_rowconfigure(0, weight=1)
        root.protocol("WM_DELETE_WINDOW", self._ao_fechar)

        self._configurar_estilo()
        self._montar_layout()
        self._atualizar_ui()
        root.mainloop()

    def _configurar_estilo(self):
        style = self.ttk.Style(self.root)
        try:
            style.theme_use("clam")
        except Exception:
            pass

        style.configure(
            "Monitor.Treeview",
            rowheight=28,
            font=("Segoe UI", 9),
            foreground="#173225",
            background="#fbfcf8",
            fieldbackground="#fbfcf8",
            borderwidth=0,
            relief="flat",
        )
        style.configure(
            "Monitor.Treeview.Heading",
            font=("Segoe UI Semibold", 9),
            foreground="#647562",
            background="#f2f6eb",
            relief="flat",
            borderwidth=0,
        )
        style.map(
            "Monitor.Treeview",
            background=[("selected", "#dfead5")],
            foreground=[("selected", "#173225")],
        )
        style.map("Monitor.Treeview.Heading", background=[("active", "#e7eedc")])

    def _montar_layout(self):
        shell = self.tk.Frame(self.root, bg="#f4f1e6")
        shell.grid(row=0, column=0, sticky="nsew")
        shell.grid_columnconfigure(0, weight=1)
        shell.grid_rowconfigure(2, weight=1)

        hero = self.tk.Frame(shell, bg="#f4f1e6")
        hero.grid(row=0, column=0, sticky="ew", padx=20, pady=(16, 10))
        hero.grid_columnconfigure(0, weight=19, uniform="hero")
        hero.grid_columnconfigure(1, weight=9, uniform="hero")

        hero_main = self._criar_painel(hero, "#fcfdf9", "#d8e2d0")
        hero_main.grid(row=0, column=0, sticky="nsew", padx=(0, 8))
        hero_main.grid_columnconfigure(0, weight=1)

        faixa = self.tk.Frame(hero_main, bg="#6d9444", height=58)
        faixa.grid(row=0, column=0, sticky="ew")
        faixa.grid_propagate(False)
        self.tk.Label(
            faixa,
            text="MONITOR OPERACIONAL",
            bg="#6d9444",
            fg="#f5faec",
            font=("Segoe UI Semibold", 9),
            padx=14,
            pady=4,
        ).pack(anchor="w", padx=18, pady=(8, 0))

        conteudo = self.tk.Frame(hero_main, bg="#fcfdf9")
        conteudo.grid(row=1, column=0, sticky="nsew", padx=22, pady=(12, 14))
        conteudo.grid_columnconfigure(0, weight=1)

        self.tk.Label(
            conteudo,
            text="Central Operacional de Lançamentos | Toll Coonagro",
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 24),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")

        self.tk.Label(
            conteudo,
            text="Monitoramento de notas de industrialização, com motor integrado para lançamento em Excel via SAP GUI.",
            bg="#fcfdf9",
            fg="#5d725f",
            font=("Segoe UI", 10),
            justify="left",
            wraplength=760,
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(4, 10))

        status_row = self.tk.Frame(conteudo, bg="#fcfdf9")
        status_row.grid(row=2, column=0, sticky="ew")
        status_row.grid_columnconfigure(1, weight=1)

        status_pill = self._criar_painel(status_row, "#e9f1df", "#d7e2cd")
        status_pill.grid(row=0, column=0, sticky="w")
        status_pill.grid_columnconfigure(1, weight=1)

        self._status_dot = self.tk.Canvas(
            status_pill,
            width=16,
            height=16,
            bg="#e9f1df",
            highlightthickness=0,
            bd=0,
        )
        self._status_dot.grid(row=0, column=0, padx=(12, 8), pady=10)
        self._status_dot_item = self._status_dot.create_oval(
            3,
            3,
            13,
            13,
            fill="#5da94d",
            outline="#5da94d",
        )

        self._label_status = self.tk.Label(
            status_pill,
            text="Atualizando painel...",
            bg="#e9f1df",
            fg="#2e5030",
            font=("Segoe UI Semibold", 11),
            anchor="w",
        )
        self._label_status.grid(row=0, column=1, sticky="w", padx=(0, 14), pady=8)

        self._label_updated = self.tk.Label(
            status_row,
            text="Atualizado em -",
            bg="#fcfdf9",
            fg="#6e816a",
            font=("Segoe UI", 10),
            anchor="w",
        )
        self._label_updated.grid(row=0, column=1, sticky="w", padx=(12, 0))

        resumo = self.tk.Frame(conteudo, bg="#fcfdf9")
        resumo.grid(row=3, column=0, sticky="w", pady=(8, 0))
        self._summary_labels["grupos_pendentes_topo"] = self._criar_resumo_pill(
            resumo, 0, "Pendentes", "0 grupos"
        )
        self._summary_labels["extraidas_hoje_topo"] = self._criar_resumo_pill(
            resumo, 1, "Hoje", "0 grupos"
        )

        hero_side = self._criar_painel(hero, "#f9fbf4", "#d8e2d0")
        hero_side.grid(row=0, column=1, sticky="new", padx=(8, 0))
        hero_side.grid_columnconfigure(0, weight=1)

        acoes = self.tk.Frame(hero_side, bg="#f9fbf4")
        acoes.grid(row=0, column=0, sticky="ew", padx=12, pady=12)
        for col in range(2):
            acoes.grid_columnconfigure(col, weight=1)

        self._criar_botao_acao(
            acoes,
            0,
            "Abrir Base",
            self._abrir_xlsm,
            "#648e43",
            "#f7fbf2",
            "#5b8140",
        )
        self._criar_botao_acao(
            acoes,
            1,
            "Atualizar",
            self._atualizar_ui,
            "#edf4e5",
            "#43652d",
            "#e2ecda",
        )

        bloco_logo = self._criar_painel(hero_side, "#ffffff", "#d8e2d0")
        bloco_logo.grid(row=1, column=0, sticky="ew", padx=12, pady=(0, 10))
        bloco_logo.grid_columnconfigure(0, weight=1)
        self._montar_bloco_logo(bloco_logo)

        bloco_industrializacao = self._criar_painel(hero_side, "#ffffff", "#d8e2d0")
        bloco_industrializacao.grid(row=2, column=0, sticky="ew", padx=12, pady=(0, 12))
        bloco_industrializacao.grid_columnconfigure(0, weight=1)
        self._montar_bloco_industrializacao(bloco_industrializacao)

        cards = self.tk.Frame(shell, bg="#f4f1e6")
        cards.grid(row=1, column=0, sticky="ew", padx=20, pady=(0, 10))
        total_cards = len(self.CARD_SPECS)
        if total_cards == 5:
            for col in range(5):
                cards.grid_columnconfigure(col, weight=1, uniform="cards")
            posicoes_cards = tuple((0, idx, 1) for idx in range(total_cards))
        else:
            for col in range(3):
                cards.grid_columnconfigure(col, weight=1, uniform="cards")
            posicoes_cards = tuple((idx // 3, idx % 3, 1) for idx in range(total_cards))

        for (chave, titulo, selo, cor), (linha, coluna, colspan) in zip(
            self.CARD_SPECS, posicoes_cards
        ):
            self._criar_cartao_metrica(
                cards, linha, coluna, chave, titulo, selo, cor, colspan=colspan
            )

        workspace = self.tk.Frame(shell, bg="#f4f1e6")
        workspace.grid(row=2, column=0, sticky="nsew", padx=20, pady=(0, 20))
        workspace.grid_columnconfigure(0, weight=1)
        workspace.grid_rowconfigure(0, weight=1)

        historico = self._criar_painel(workspace, "#fcfdf9", "#d8e2d0")
        historico.grid(row=0, column=0, sticky="nsew")
        historico.grid_columnconfigure(0, weight=1)
        historico.grid_rowconfigure(1, weight=1)

        topo_hist = self.tk.Frame(historico, bg="#fcfdf9")
        topo_hist.grid(row=0, column=0, sticky="ew", padx=16, pady=(16, 8))
        topo_hist.grid_columnconfigure(0, weight=1)

        self.tk.Label(
            topo_hist,
            text="Acompanhamento das Notas",
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 18),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        self.tk.Label(
            topo_hist,
            text=f"Histórico dos últimos {JANELA_HISTORICO_DIAS} dias. Selecione uma 5124 para abrir NF PAR, materiais e volumes vinculados.",
            bg="#fcfdf9",
            fg="#6e816a",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=1, column=0, sticky="w", pady=(6, 0))

        acoes_hist = self.tk.Frame(topo_hist, bg="#fcfdf9")
        acoes_hist.grid(row=0, column=1, rowspan=2, sticky="e")
        self._label_history_meta = self.tk.Label(
            acoes_hist,
            text=f"0 registros | {JANELA_HISTORICO_DIAS} dias",
            bg="#fcfdf9",
            fg="#6e816a",
            font=("Segoe UI", 10),
            anchor="e",
        )
        self._label_history_meta.grid(row=0, column=0, sticky="e", padx=(0, 10))

        self.tk.Button(
            acoes_hist,
            text="Detalhes da seleção",
            command=self._abrir_detalhes_historico,
            bg="#edf4e5",
            fg="#43652d",
            activebackground="#e2ecda",
            activeforeground="#43652d",
            relief="flat",
            bd=0,
            padx=12,
            pady=7,
            font=("Segoe UI Semibold", 9),
            cursor="hand2",
        ).grid(row=0, column=1, sticky="e")

        tabela_borda = self._criar_painel(historico, "#ffffff", "#e4ead9")
        tabela_borda.grid(row=1, column=0, sticky="nsew", padx=16, pady=(0, 16))
        tabela_borda.grid_columnconfigure(0, weight=1)
        tabela_borda.grid_rowconfigure(0, weight=1)

        colunas = ("nf", "nf_par", "emissao", "toneladas", "status", "entrada")
        tree = self.ttk.Treeview(
            tabela_borda,
            columns=colunas,
            show="headings",
            selectmode="browse",
            style="Monitor.Treeview",
        )
        tree.heading("nf", text="NF 5124")
        tree.heading("nf_par", text="NF PAR")
        tree.heading("emissao", text="Emissão")
        tree.heading("toneladas", text="Toneladas")
        tree.heading("status", text="Status")
        tree.heading("entrada", text="Extraída em")
        tree.column("nf", width=105, anchor="center")
        tree.column("nf_par", width=125, anchor="center")
        tree.column("emissao", width=110, anchor="center")
        tree.column("toneladas", width=120, anchor="e")
        tree.column("status", width=130, anchor="center")
        tree.column("entrada", width=150, anchor="center")
        tree.grid(row=0, column=0, sticky="nsew")
        tree.bind("<Double-1>", self._abrir_detalhes_historico)
        tree.bind("<<TreeviewSelect>>", self._atualizar_detalhes_historico_se_visivel)

        scroll = self.ttk.Scrollbar(tabela_borda, orient="vertical", command=tree.yview)
        scroll.grid(row=0, column=1, sticky="ns")
        tree.configure(yscrollcommand=scroll.set)
        tree.tag_configure("pendente", background="#f7ecd4")
        tree.tag_configure("lancada_hoje", background="#def0e5")
        tree.tag_configure("lancada", background="#ecf4e7")
        self._tree = tree

        detalhes = self._criar_painel(historico, "#fcfdf9", "#e4ead9")
        detalhes.grid(row=2, column=0, sticky="ew", padx=16, pady=(0, 16))
        detalhes.grid_columnconfigure(0, weight=1)
        detalhes.grid_rowconfigure(1, weight=1)

        topo_detalhes = self.tk.Frame(detalhes, bg="#fcfdf9")
        topo_detalhes.grid(row=0, column=0, sticky="ew", padx=14, pady=(12, 8))
        topo_detalhes.grid_columnconfigure(0, weight=1)

        self._details_title = self.tk.Label(
            topo_detalhes,
            text="Materiais vinculados",
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 14),
            anchor="w",
        )
        self._details_title.grid(row=0, column=0, sticky="w")

        self.tk.Button(
            topo_detalhes,
            text="Ocultar",
            command=self._ocultar_detalhes_historico,
            bg="#f3f5ef",
            fg="#5d725f",
            activebackground="#e6eadf",
            activeforeground="#5d725f",
            relief="flat",
            bd=0,
            padx=10,
            pady=6,
            font=("Segoe UI Semibold", 9),
            cursor="hand2",
        ).grid(row=0, column=1, sticky="e")

        self._details_meta = self.tk.Label(
            topo_detalhes,
            text="",
            bg="#fcfdf9",
            fg="#6e816a",
            font=("Segoe UI", 9),
            anchor="w",
        )
        self._details_meta.grid(row=1, column=0, columnspan=2, sticky="w", pady=(6, 0))

        cards_detalhes = self._criar_painel(detalhes, "#ffffff", "#e4ead9")
        cards_detalhes.grid(row=1, column=0, sticky="nsew", padx=14, pady=(0, 14))
        cards_detalhes.grid_columnconfigure(0, weight=1)
        cards_detalhes.grid_rowconfigure(0, weight=1)

        canvas_detalhes = self.tk.Canvas(
            cards_detalhes,
            bg="#ffffff",
            highlightthickness=0,
            bd=0,
            height=260,
        )
        canvas_detalhes.grid(row=0, column=0, sticky="nsew")

        scroll_detalhes = self.ttk.Scrollbar(
            cards_detalhes, orient="vertical", command=canvas_detalhes.yview
        )
        scroll_detalhes.grid(row=0, column=1, sticky="ns")
        canvas_detalhes.configure(yscrollcommand=scroll_detalhes.set)

        cards_frame = self.tk.Frame(canvas_detalhes, bg="#ffffff")
        cards_frame.grid_columnconfigure(0, weight=1, uniform="detalhes")
        cards_frame.grid_columnconfigure(1, weight=1, uniform="detalhes")
        details_window = canvas_detalhes.create_window(
            (0, 0), window=cards_frame, anchor="nw"
        )
        cards_frame.bind("<Configure>", self._sincronizar_scroll_detalhes)
        canvas_detalhes.bind("<Configure>", self._sincronizar_scroll_detalhes)

        self._details_cards_canvas = canvas_detalhes
        self._details_cards_frame = cards_frame
        self._details_cards_window = details_window
        self._details_panel = detalhes
        self._details_panel.grid_remove()

    def _criar_painel(self, parent, bg, border):
        return self.tk.Frame(
            parent,
            bg=bg,
            bd=0,
            highlightthickness=1,
            highlightbackground=border,
            highlightcolor=border,
        )

    def _criar_botao_acao(self, parent, coluna, texto, comando, bg, fg, active_bg):
        botao = self.tk.Button(
            parent,
            text=texto,
            command=comando,
            bg=bg,
            fg=fg,
            activebackground=active_bg,
            activeforeground=fg,
            relief="flat",
            bd=0,
            padx=12,
            pady=9,
            font=("Segoe UI Semibold", 10),
            cursor="hand2",
        )
        botao.grid(row=0, column=coluna, sticky="ew", padx=(0, 8) if coluna == 0 else 0)

    def _criar_resumo_pill(self, parent, coluna, titulo, valor):
        pill = self._criar_painel(parent, "#ffffff", "#d8e2d0")
        pill.grid(row=0, column=coluna, sticky="w", padx=(0, 6), pady=0)
        self.tk.Label(
            pill,
            text=titulo,
            bg="#ffffff",
            fg="#5d725f",
            font=("Segoe UI", 9),
        ).grid(row=0, column=0, sticky="w", padx=(12, 8), pady=7)
        lbl = self.tk.Label(
            pill,
            text=valor,
            bg="#ffffff",
            fg="#173225",
            font=("Segoe UI Semibold", 10),
        )
        lbl.grid(row=0, column=1, sticky="w", padx=(0, 12), pady=7)
        return lbl

    def _caminhos_logo_monitor(self) -> list[str]:
        base_dir = _diretorio_aplicacao()
        assets_dir = os.path.join(base_dir, "assets")
        candidatos = [
            os.path.join(base_dir, "assets", "logo.png"),
            os.path.join(base_dir, "assets", "logo_monitor.png"),
            os.path.join(base_dir, "assets", "coonagro_logo.png"),
            os.path.join(base_dir, "assets", "coonagro.png"),
            os.path.join(base_dir, "assets", "mosaic_R_blk_5c_rgb.png"),
        ]
        vistos = {os.path.normcase(caminho) for caminho in candidatos}

        if os.path.isdir(assets_dir):
            for nome_arquivo in os.listdir(assets_dir):
                nome_normalizado = nome_arquivo.lower()
                if not nome_normalizado.endswith((".png", ".gif", ".pgm", ".ppm")):
                    continue
                if not any(
                    termo in nome_normalizado
                    for termo in ("logo", "mosaic", "coonagro")
                ):
                    continue
                caminho = os.path.join(assets_dir, nome_arquivo)
                chave = os.path.normcase(caminho)
                if chave in vistos:
                    continue
                candidatos.append(caminho)
                vistos.add(chave)

        return candidatos

    def _carregar_logo_monitor(self, largura_max: int = 220, altura_max: int = 88):
        for caminho in self._caminhos_logo_monitor():
            if not os.path.exists(caminho):
                continue
            try:
                imagem = self.tk.PhotoImage(file=caminho)
                largura = max(imagem.width(), 1)
                altura = max(imagem.height(), 1)
                fator = max(
                    (largura + largura_max - 1) // largura_max,
                    (altura + altura_max - 1) // altura_max,
                    1,
                )
                if fator > 1:
                    imagem = imagem.subsample(fator, fator)
                self._logo_photo = imagem
                return imagem
            except Exception:
                continue
        self._logo_photo = None
        return None

    def _desenhar_logo_fallback(self, parent):
        canvas = self.tk.Canvas(
            parent,
            width=220,
            height=88,
            bg="#ffffff",
            highlightthickness=0,
            bd=0,
        )
        canvas.grid(row=0, column=0, sticky="ew", padx=12, pady=10)
        canvas.create_rectangle(8, 8, 212, 80, fill="#f6f9ef", outline="#dfe7d4")
        canvas.create_oval(20, 20, 62, 62, fill="#6d9444", outline="#6d9444")
        canvas.create_oval(40, 40, 70, 70, fill="#d39b2d", outline="#d39b2d")
        canvas.create_polygon(
            30,
            56,
            48,
            18,
            58,
            28,
            42,
            64,
            fill="#edf4e5",
            outline="#edf4e5",
        )
        canvas.create_text(
            84,
            34,
            text="COONAGRO",
            anchor="w",
            fill="#173225",
            font=("Segoe UI Semibold", 14),
        )
        canvas.create_text(
            84,
            54,
            text="Painel Operacional",
            anchor="w",
            fill="#5d725f",
            font=("Segoe UI", 9),
        )
        return canvas

    def _montar_bloco_logo(self, parent):
        imagem_logo = self._carregar_logo_monitor()
        if imagem_logo is not None:
            self.tk.Label(
                parent,
                image=imagem_logo,
                bg="#ffffff",
                anchor="center",
            ).grid(row=0, column=0, sticky="ew", padx=12, pady=10)
        else:
            self._desenhar_logo_fallback(parent)

    def _montar_bloco_industrializacao(self, parent):
        parent.grid_columnconfigure(0, weight=1, uniform="industrializacao")
        parent.grid_columnconfigure(1, weight=1, uniform="industrializacao")
        self.tk.Label(
            parent,
            text="Lançadas hoje por industrialização",
            bg="#ffffff",
            fg="#173225",
            font=("Segoe UI Semibold", 10),
            anchor="w",
        ).grid(row=0, column=0, columnspan=2, sticky="w", padx=14, pady=(10, 2))

        self._label_industrializacao_meta = self.tk.Label(
            parent,
            text="Nenhum grupo lançado hoje",
            bg="#ffffff",
            fg="#6e816a",
            font=("Segoe UI", 8),
            anchor="w",
        )
        self._label_industrializacao_meta.grid(
            row=1, column=0, columnspan=2, sticky="w", padx=14, pady=(0, 8)
        )

        self._industrializacao_labels = {}
        for indice, (chave, _, titulo, cor) in enumerate(
            INDUSTRIALIZACAO_MONITOR_SPECS
        ):
            linha_idx = 2 + (indice // 2)
            coluna_idx = indice % 2
            linha = self._criar_painel(parent, "#fcfdf9", "#e4ead9")
            linha.grid(
                row=linha_idx,
                column=coluna_idx,
                sticky="ew",
                padx=(12, 6) if coluna_idx == 0 else (6, 12),
                pady=(0, 8),
            )
            linha.grid_columnconfigure(1, weight=1)

            marcador = self.tk.Canvas(
                linha,
                width=12,
                height=12,
                bg="#fcfdf9",
                highlightthickness=0,
                bd=0,
            )
            marcador.grid(row=0, column=0, padx=(10, 6), pady=8)
            marcador.create_oval(2, 2, 10, 10, fill=cor, outline=cor)

            self.tk.Label(
                linha,
                text=titulo,
                bg="#fcfdf9",
                fg="#173225",
                font=("Segoe UI", 8),
                anchor="w",
            ).grid(row=0, column=1, sticky="w", pady=7)

            valor = self.tk.Label(
                linha,
                text="0",
                bg="#fcfdf9",
                fg="#173225",
                font=("Segoe UI Semibold", 11),
                anchor="e",
            )
            valor.grid(row=0, column=2, sticky="e", padx=(6, 10), pady=7)
            self._industrializacao_labels[chave] = valor

    def _criar_cartao_metrica(
        self, parent, linha, coluna, chave, titulo, selo, cor, colspan=1
    ):
        card = self._criar_painel(parent, "#fcfdf9", "#d8e2d0")
        card.grid(
            row=linha,
            column=coluna,
            columnspan=colspan,
            sticky="nsew",
            padx=5,
            pady=4,
        )
        card.grid_columnconfigure(1, weight=1)

        self.tk.Frame(card, bg=cor, width=7).grid(
            row=0, column=0, rowspan=2, sticky="ns"
        )

        topo = self.tk.Frame(card, bg="#fcfdf9")
        topo.grid(row=0, column=1, sticky="ew", padx=10, pady=(6, 2))
        topo.grid_columnconfigure(0, weight=1)

        self.tk.Label(
            topo,
            text=titulo,
            bg="#fcfdf9",
            fg="#5d725f",
            font=("Segoe UI Semibold", 9),
        ).grid(row=0, column=0, sticky="w")
        self.tk.Label(
            topo,
            text=selo,
            bg="#ffffff",
            fg=cor,
            font=("Segoe UI Semibold", 9),
            padx=8,
            pady=3,
        ).grid(row=0, column=1, sticky="e")

        lbl = self.tk.Label(
            card,
            text="-",
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 15),
            anchor="w",
        )
        lbl.grid(row=1, column=1, sticky="w", padx=10, pady=(0, 6))
        self._labels_cards[chave] = lbl

    def _criar_item_hoje(self, parent, linha, titulo, valor):
        item = self._criar_painel(parent, "#f6f1df", "#ece0bf")
        item.grid(row=linha, column=0, sticky="ew", pady=(0, 8 if linha < 2 else 0))
        self.tk.Label(
            item,
            text=titulo,
            bg="#f6f1df",
            fg="#173225",
            font=("Segoe UI Semibold", 10),
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=12, pady=(10, 4))
        lbl = self.tk.Label(
            item,
            text=valor,
            bg="#f6f1df",
            fg="#5d725f",
            font=("Segoe UI", 10),
            anchor="w",
        )
        lbl.grid(row=1, column=0, sticky="w", padx=12, pady=(0, 10))
        return lbl

    def _criar_bloco_lateral(self, parent, titulo, subtitulo, itens):
        self.tk.Label(
            parent,
            text=titulo,
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 18),
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=20, pady=(20, 4))
        self.tk.Label(
            parent,
            text=subtitulo,
            bg="#fcfdf9",
            fg="#5d725f",
            font=("Segoe UI", 10),
            justify="left",
            wraplength=320,
            anchor="w",
        ).grid(row=1, column=0, sticky="w", padx=20, pady=(0, 14))

        lista = self.tk.Frame(parent, bg="#fcfdf9")
        lista.grid(row=2, column=0, sticky="ew", padx=20, pady=(0, 20))
        lista.grid_columnconfigure(0, weight=1)

        for idx, (cabecalho, texto) in enumerate(itens):
            item = self._criar_painel(lista, "#f6f1df", "#ece0bf")
            item.grid(
                row=idx,
                column=0,
                sticky="ew",
                pady=(0, 10 if idx < len(itens) - 1 else 0),
            )
            self.tk.Label(
                item,
                text=cabecalho,
                bg="#f6f1df",
                fg="#173225",
                font=("Segoe UI Semibold", 10),
                anchor="w",
            ).grid(row=0, column=0, sticky="w", padx=14, pady=(12, 4))
            self.tk.Label(
                item,
                text=texto,
                bg="#f6f1df",
                fg="#5d725f",
                font=("Segoe UI", 10),
                justify="left",
                wraplength=300,
                anchor="w",
            ).grid(row=1, column=0, sticky="w", padx=14, pady=(0, 12))

    def _criar_legenda(self, parent):
        self.tk.Label(
            parent,
            text="Status",
            bg="#fcfdf9",
            fg="#173225",
            font=("Segoe UI Semibold", 18),
            anchor="w",
        ).grid(row=0, column=0, sticky="w", padx=18, pady=(18, 10))

        itens = [
            ("Pendente", "#f7ecd4", "#8a6113"),
            ("Lançada", "#e7f2e3", "#43652d"),
            ("Lançada hoje", "#def0e5", "#1b6b58"),
        ]

        lista = self.tk.Frame(parent, bg="#fcfdf9")
        lista.grid(row=1, column=0, sticky="w", padx=18, pady=(0, 18))

        for idx, (titulo, bg, fg) in enumerate(itens):
            badge = self.tk.Label(
                lista,
                text=titulo,
                bg=bg,
                fg=fg,
                font=("Segoe UI Semibold", 9),
                padx=12,
                pady=6,
            )
            badge.grid(
                row=0,
                column=idx,
                sticky="w",
                padx=(0, 8 if idx < len(itens) - 1 else 0),
            )

    def _texto_status_base(self, resumo: dict) -> str:
        status = str(resumo.get("status_monitor") or "")
        status_upper = status.upper()

        if "FALHA" in status_upper or "ERRO" in status_upper:
            return "Falha operacional"
        if "FECHE O XLSM" in status_upper:
            return "Fechar XLSM"
        if "ATUALIZANDO BASE" in status_upper:
            return "Atualizando agora"
        if "BASE ATUALIZADA" in status_upper:
            return "Base atualizada"
        if "SAP" in status_upper:
            return "Aguardando SAP"
        if int(resumo.get("novas_notas") or 0) > 0:
            return "Fila aguardando recarga"
        return "Automática ativa"

    def _texto_status_painel(self, resumo: dict) -> str:
        status = str(resumo.get("status_monitor") or "")
        status_upper = status.upper()

        if "FALHA" in status_upper or "ERRO" in status_upper:
            return "Falha ao atualizar painel"
        pendentes = int(resumo.get("grupos_pendentes") or 0)
        if pendentes <= 0:
            return "Sem lançamentos pendentes"
        return f"Aguardando lancamento - {self._texto_grupos(pendentes)}"

    def _texto_subtitulo_painel(self, resumo: dict) -> str:
        status = str(resumo.get("status_monitor") or "")
        status_upper = status.upper()

        if "FALHA" in status_upper or "ERRO" in status_upper:
            return "O monitor encontrou uma falha ao montar os dados do painel."
        if "FECHE O XLSM" in status_upper:
            return "Feche o XLSM para permitir a atualização automática da Base."
        if int(resumo.get("novas_notas") or 0) > 0:
            return "Existem XMLs válidos aguardando atualização da Base."
        return "Sem pendências relevantes no fluxo neste momento."

    def _texto_fila_recarrega(self, resumo: dict) -> str:
        return self._texto_xml(int(resumo.get("novas_notas") or 0))

    def _texto_grupos(self, valor) -> str:
        qtd = int(valor or 0)
        return f"{qtd} grupo" if qtd == 1 else f"{qtd} grupos"

    def _texto_xml(self, valor) -> str:
        qtd = int(valor or 0)
        return f"{qtd} XML"

    def _texto_valor_cartao(self, chave: str, resumo: dict) -> str:
        if chave in {"novas_notas"}:
            return self._texto_xml(resumo.get(chave) or 0)
        if chave in {
            "grupos_pendentes",
            "grupos_lancados_hoje",
            "grupos_extraidos_hoje",
        }:
            return self._texto_grupos(resumo.get(chave) or 0)
        return _formatar_toneladas(resumo.get(chave) or 0)

    def _texto_nota_metrica(self, chave: str, resumo: dict) -> str:
        if chave == "novas_notas":
            if int(resumo.get("novas_notas") or 0) > 0:
                return "XMLs válidos aguardando atualização da Base."
            return "Nenhum XML pendente para recarregar agora."

        if chave == "grupos_pendentes":
            if int(resumo.get("grupos_pendentes") or 0) > 0:
                return "Volume operacional ainda aguardando fechamento."
            return "Nenhum grupo pendente no momento."

        if chave == "toneladas_pendentes":
            if float(resumo.get("toneladas_pendentes") or 0) > 0:
                return "Indicador direto da carga parada no banco."
            return "Sem toneladas aguardando processamento."

        if chave == "grupos_lancados_hoje":
            if int(resumo.get("grupos_lancados_hoje") or 0) > 0:
                return "Grupos já concluídos no fechamento do dia."
            return "Nenhum grupo lançado hoje até agora."

        if chave == "toneladas_lancadas_hoje":
            if float(resumo.get("toneladas_lancadas_hoje") or 0) > 0:
                return "Volume total já consolidado na operação de hoje."
            return "Sem toneladas lançadas hoje até agora."

        if resumo.get("grupos_extraidos_hoje") is not None:
            return "NFs 5124 efetivamente capturadas no dia."
        return "Ainda sem leitura consolidada de extração no dia."

    def _atualizar_bloco_industrializacao(self, resumo: dict):
        contagens = resumo.get("industrializacoes_lancadas_hoje") or {}
        total_lancadas = int(resumo.get("grupos_lancados_hoje") or 0)
        sem_tipo = int(resumo.get("grupos_lancados_hoje_sem_tipo") or 0)

        if self._label_industrializacao_meta is not None:
            if total_lancadas <= 0:
                texto_meta = "Nenhum grupo lançado hoje"
            elif sem_tipo > 0:
                texto_meta = f"{self._texto_grupos(total_lancadas)} hoje | {sem_tipo} sem tipo mapeado"
            else:
                texto_meta = f"{self._texto_grupos(total_lancadas)} hoje"
            self._label_industrializacao_meta.configure(text=texto_meta)

        for chave, label in self._industrializacao_labels.items():
            label.configure(text=str(int(contagens.get(chave) or 0)))

    def _atualizar_cor_status(self, resumo: dict):
        if self._status_dot is None or self._status_dot_item is None:
            return

        status = str(resumo.get("status_monitor") or "").upper()
        cor_fundo = "#e9f1df"
        cor = "#5da94d"

        if "FALHA" in status or "ERRO" in status:
            cor_fundo = "#f7e1dc"
            cor = "#c84f42"
        elif "FECHE O XLSM" in status or int(resumo.get("novas_notas") or 0) > 0:
            cor_fundo = "#f7ecd4"
            cor = "#d39b2d"

        self._status_dot.configure(bg=cor_fundo)
        self._status_dot.itemconfig(self._status_dot_item, fill=cor, outline=cor)

    def _ao_fechar(self):
        global _painel_monitor_ref
        self._fechado = True
        try:
            if self.root is not None:
                if self._after_id is not None:
                    try:
                        self.root.after_cancel(self._after_id)
                    except Exception:
                        pass
                    self._after_id = None
                self.root.destroy()
        finally:
            self.root = None
            with _painel_monitor_lock:
                if _painel_monitor_ref is self:
                    _painel_monitor_ref = None

    def _abrir_xlsm(self):
        if not _abrir_base_xlsm():
            self.messagebox.showwarning(
                "Base não encontrada",
                f"Não foi possível localizar o arquivo:\n{caminho_excel}",
            )

    def _abrir_log(self):
        if not _abrir_log_monitor():
            self.messagebox.showwarning(
                "Log não encontrado",
                f"Não foi possível localizar o log:\n{_log_path}",
            )

    def _obter_historico_selecionado(self):
        if self._tree is None:
            return None, None
        selecao = self._tree.selection()
        if not selecao:
            return None, None
        iid = selecao[0]
        return iid, self._history_rows.get(iid)

    def _sincronizar_scroll_detalhes(self, event=None):
        if self._details_cards_canvas is None or self._details_cards_frame is None:
            return
        self._details_cards_canvas.configure(
            scrollregion=self._details_cards_canvas.bbox("all")
        )
        if self._details_cards_window is not None:
            largura = max(self._details_cards_canvas.winfo_width(), 1)
            self._details_cards_canvas.itemconfigure(
                self._details_cards_window, width=largura
            )

    def _limpar_cards_detalhes(self):
        if self._details_cards_frame is None:
            return
        for child in self._details_cards_frame.winfo_children():
            child.destroy()

    def _criar_card_detalhe_material(self, parent, detalhe: dict, indice: int):
        cor_destaque = "#648e43" if indice % 2 == 0 else "#d39b2d"

        card = self._criar_painel(parent, "#ffffff", "#e4ead9")
        card.grid(
            row=indice // 2,
            column=indice % 2,
            sticky="nsew",
            padx=6,
            pady=6,
        )
        card.grid_columnconfigure(1, weight=1)

        self.tk.Frame(card, bg=cor_destaque, width=6).grid(
            row=0, column=0, rowspan=3, sticky="ns"
        )

        topo = self.tk.Frame(card, bg="#ffffff")
        topo.grid(row=0, column=1, sticky="ew", padx=14, pady=(12, 8))
        topo.grid_columnconfigure(0, weight=1)

        self.tk.Label(
            topo,
            text=detalhe.get("cod_material") or "Material",
            bg="#ffffff",
            fg="#43652d",
            font=("Segoe UI Semibold", 9),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        self.tk.Label(
            topo,
            text=f"NF PAR {detalhe.get('nf_par') or '-'}",
            bg="#eef4e7",
            fg="#43652d",
            font=("Segoe UI Semibold", 8),
            padx=8,
            pady=4,
        ).grid(row=0, column=1, sticky="e")

        self.tk.Label(
            card,
            text=detalhe.get("descricao") or "Material sem descrição",
            bg="#ffffff",
            fg="#173225",
            font=("Segoe UI Semibold", 10),
            justify="left",
            wraplength=320,
            anchor="w",
        ).grid(row=1, column=1, sticky="w", padx=14)

        rodape = self.tk.Frame(card, bg="#ffffff")
        rodape.grid(row=2, column=1, sticky="ew", padx=14, pady=(10, 12))
        rodape.grid_columnconfigure(1, weight=1)
        self.tk.Label(
            rodape,
            text="Volume detalhado",
            bg="#ffffff",
            fg="#6e816a",
            font=("Segoe UI", 9),
            anchor="w",
        ).grid(row=0, column=0, sticky="w")
        self.tk.Label(
            rodape,
            text=f"{_formatar_toneladas(detalhe.get('toneladas') or 0)} {detalhe.get('un') or ''}",
            bg="#ffffff",
            fg="#173225",
            font=("Segoe UI Semibold", 11),
            anchor="e",
        ).grid(row=0, column=1, sticky="e")

    def _ocultar_detalhes_historico(self):
        if self._details_panel is not None:
            self._details_panel.grid_remove()
        self._limpar_cards_detalhes()
        self._details_current_iid = None

    def _preencher_detalhes_historico(self, iid: str, item: dict | None):
        if (
            item is None
            or self._details_panel is None
            or self._details_cards_frame is None
            or self._details_title is None
            or self._details_meta is None
        ):
            return

        self._details_current_iid = iid
        self._details_title.configure(
            text=f"Materiais vinculados a NF 5124 {item['nf']}"
        )

        detalhes = item.get("detalhes") or []
        nf_par = item.get("nf_par") or "-"
        self._details_meta.configure(
            text=f"NF PAR: {nf_par} | {len(detalhes)} material(is) vinculado(s)"
        )

        self._limpar_cards_detalhes()
        if detalhes:
            for indice, detalhe in enumerate(detalhes):
                self._criar_card_detalhe_material(
                    self._details_cards_frame, detalhe, indice
                )
        else:
            vazio = self._criar_painel(self._details_cards_frame, "#fbfcf8", "#e4ead9")
            vazio.grid(row=0, column=0, columnspan=2, sticky="ew", padx=6, pady=6)
            self.tk.Label(
                vazio,
                text="Sem materiais vinculados a esta NF.",
                bg="#fbfcf8",
                fg="#6e816a",
                font=("Segoe UI", 10),
                anchor="center",
                pady=16,
            ).grid(row=0, column=0, sticky="ew", padx=12)

        self._sincronizar_scroll_detalhes()
        self._details_panel.grid()

    def _abrir_detalhes_historico(self, event=None):
        iid, item = self._obter_historico_selecionado()
        if iid is None or item is None:
            if event is None and self.messagebox is not None:
                self.messagebox.showinfo(
                    "Selecione uma nota",
                    "Selecione uma nota do histórico para ver os materiais e volumes detalhados.",
                )
            return

        self._preencher_detalhes_historico(iid, item)

    def _atualizar_detalhes_historico_se_visivel(self, event=None):
        if self._details_panel is None or not self._details_panel.winfo_ismapped():
            return
        iid, item = self._obter_historico_selecionado()
        if iid is None or item is None:
            return
        self._preencher_detalhes_historico(iid, item)

    def _atualizar_ui(self):
        if self._fechado or self.root is None:
            return

        if self._after_id is not None:
            try:
                self.root.after_cancel(self._after_id)
            except Exception:
                pass
            self._after_id = None

        try:
            resumo = _carregar_resumo_monitor()
            self._label_status.configure(text=self._texto_status_painel(resumo))
            self._label_updated.configure(
                text=f"Atualizado em {time.strftime('%d/%m/%Y %H:%M:%S')}"
            )
            self._atualizar_cor_status(resumo)

            self._summary_labels["grupos_pendentes_topo"].configure(
                text=self._texto_grupos(resumo.get("grupos_pendentes") or 0)
            )
            self._summary_labels["extraidas_hoje_topo"].configure(
                text=self._texto_grupos(resumo.get("grupos_extraidos_hoje") or 0)
            )

            for chave in self._labels_cards:
                self._labels_cards[chave].configure(
                    text=self._texto_valor_cartao(chave, resumo)
                )

            self._atualizar_bloco_industrializacao(resumo)

            if self._label_history_meta is not None:
                self._label_history_meta.configure(
                    text=f"{len(resumo['historico'])} registros | {JANELA_HISTORICO_DIAS} dias"
                )

            selecionado_iid = self._details_current_iid
            if selecionado_iid is None and self._tree is not None:
                selecao_atual = self._tree.selection()
                if selecao_atual:
                    selecionado_iid = selecao_atual[0]

            self._history_rows = {}
            self._tree.delete(*self._tree.get_children())
            for item in resumo["historico"]:
                tag = "pendente"
                if item["status"] == "Lançada hoje":
                    tag = "lancada_hoje"
                elif item["status"] == "Lançada":
                    tag = "lancada"

                iid = f"nf_{item['nf']}"
                self._history_rows[iid] = item

                self._tree.insert(
                    "",
                    "end",
                    iid=iid,
                    values=(
                        item["nf"],
                        item.get("nf_par") or "-",
                        _formatar_data_br(item["emissao"]),
                        _formatar_toneladas(item["toneladas"]),
                        item["status"],
                        _formatar_datahora_br(
                            item["data_importacao"] or item["emissao"]
                        ),
                    ),
                    tags=(tag,),
                )

            if selecionado_iid and self._tree.exists(selecionado_iid):
                self._tree.selection_set(selecionado_iid)
                self._tree.focus(selecionado_iid)
                if (
                    self._details_panel is not None
                    and self._details_panel.winfo_ismapped()
                ):
                    self._preencher_detalhes_historico(
                        selecionado_iid, self._history_rows.get(selecionado_iid)
                    )
            elif (
                self._details_panel is not None and self._details_panel.winfo_ismapped()
            ):
                self._ocultar_detalhes_historico()
        except Exception as e:
            _log(f"[PAINEL] Falha ao atualizar interface: {e}")
            if self._label_status is not None:
                self._label_status.configure(text="Falha ao carregar painel")
        finally:
            if not self._fechado and self.root is not None:
                self._after_id = self.root.after(self.REFRESH_MS, self._atualizar_ui)


def _abrir_painel_monitor(icon=None, item=None):
    global _painel_monitor_ref
    with _painel_monitor_lock:
        if _painel_monitor_ref is not None and _painel_monitor_ref.aberto():
            return

        painel = PainelMonitorXML()
        _painel_monitor_ref = painel

    threading.Thread(target=painel.executar, daemon=True).start()


def _wb_aberto_no_excel():
    """
    Verifica se a base XLSM alvo está aberta no Excel.
    Retorna o objeto COM do Workbook se encontrado, ou None.
    """
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.GetActiveObject("Excel.Application")
    except Exception:
        return None

    caminho_alvo = os.path.abspath(caminho_excel).lower()
    nome_alvo = os.path.basename(caminho_excel).lower()

    try:
        for wb in excel.Workbooks:
            try:
                full_name = os.path.abspath(str(wb.FullName)).lower()
            except Exception:
                full_name = ""
            try:
                nome = str(wb.Name).lower()
            except Exception:
                nome = ""

            if full_name == caminho_alvo or nome == nome_alvo:
                return wb
    except Exception:
        return None

    return None


def _abrir_base_em_excel_invisivel():
    """
    Abre a Base em uma instância oculta do Excel para atualizar via COM mesmo
    quando o arquivo estiver fechado. Isso preserva layout, imagens, shapes,
    botões, macros e personalizações visuais.
    """
    if not os.path.exists(caminho_excel):
        _log(
            "[ERRO] Base XLSM não encontrada. Atualização abortada para preservar layout e personalização."
        )
        return None, None

    excel = None
    wb = None
    try:
        pythoncom.CoInitialize()
        excel = win32com.client.DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            excel.AutomationSecurity = 3  # msoAutomationSecurityForceDisable
        except Exception:
            pass
        try:
            excel.ScreenUpdating = False
        except Exception:
            pass
        try:
            excel.EnableEvents = False
        except Exception:
            pass

        wb = excel.Workbooks.Open(
            os.path.abspath(caminho_excel),
            UpdateLinks=0,
            ReadOnly=False,
            IgnoreReadOnlyRecommended=True,
            AddToMru=False,
        )
        return excel, wb
    except Exception as e:
        _log(f"[COM] Falha ao abrir a Base em Excel oculto: {e}")
        try:
            if wb is not None:
                wb.Close(SaveChanges=False)
        except Exception:
            pass
        try:
            if excel is not None:
                excel.Quit()
        except Exception:
            pass
        return None, None


def _fechar_excel_invisivel(excel_app, wb_com):
    try:
        if wb_com is not None:
            try:
                wb_com.Close(SaveChanges=False)
            except Exception:
                pass
    finally:
        if excel_app is not None:
            try:
                excel_app.EnableEvents = True
            except Exception:
                pass
            try:
                excel_app.ScreenUpdating = True
            except Exception:
                pass
            try:
                excel_app.DisplayAlerts = False
            except Exception:
                pass
            try:
                excel_app.Quit()
            except Exception:
                pass


def _normalizar_status_dashboard(texto) -> str:
    texto = str(texto or "").strip()
    if not texto:
        return ""
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(ch for ch in texto if not unicodedata.combining(ch))
    texto = re.sub(r"\s+", " ", texto)
    return texto.upper()


def _texto_status_dashboard(wb_com) -> str:
    for nome_aba in ("Dashboard", "Base"):
        try:
            ws_com = wb_com.Sheets(nome_aba)
        except Exception:
            continue

        for celula in (
            "B2",
            "B3",
            "B4",
            "C2",
            "C3",
            "C4",
            "D2",
            "D3",
            "D4",
        ):
            try:
                valor = ws_com.Range(celula).Value
            except Exception:
                continue

            status = _normalizar_status_dashboard(valor)
            if status and any(token in status for token in STATUS_SAP_BLOQUEANTES):
                return str(valor)

    return ""


def _atualizacao_bloqueada_por_status_sap(wb_com) -> bool:
    if not BLOQUEAR_ATUALIZACAO_SE_SAP_ATIVO:
        return False

    status_txt = _normalizar_status_dashboard(_texto_status_dashboard(wb_com))
    if not status_txt:
        return False

    return any(token in status_txt for token in STATUS_SAP_BLOQUEANTES)


EXCEL_UI_PROTECTION_PASSWORD = "RPA_Coonagro_UI"
EXCEL_NUMBER_FORMATS_BY_HEADER = {
    "Vlr. Unitário": ("#,##0.000000", "#.##0,000000"),
    "Vlr.Total Item": ("#,##0.00", "#.##0,00"),
    "Vlr.Total (NF)": ("#,##0.00", "#.##0,00"),
}


def _desproteger_base_operacional_com(ws_com):
    protegido = False
    try:
        protegido = bool(ws_com.ProtectContents)
    except Exception:
        protegido = False

    if protegido:
        ws_com.Unprotect(EXCEL_UI_PROTECTION_PASSWORD)

    return protegido


def _proteger_base_operacional_com(ws_com):
    ws_com.Protect(
        Password=EXCEL_UI_PROTECTION_PASSWORD,
        DrawingObjects=True,
        Contents=True,
        Scenarios=True,
        UserInterfaceOnly=True,
        AllowFiltering=True,
        AllowSorting=True,
    )


def _valor_excel_limpo(valor, como_texto: bool = False):
    if valor is None:
        return ""

    try:
        if pd.isna(valor):
            return ""
    except Exception:
        pass

    if isinstance(valor, str):
        texto = valor.strip()
        if texto.lower() in {"nan", "none", "nat", "<na>", "n/a"}:
            return ""
        return texto

    return str(valor) if como_texto else valor


def _aplicar_formatos_base_com(ws_com, header_names, first_row: int, last_row: int):
    if last_row < first_row:
        return

    for idx, col_name in enumerate(header_names, start=1):
        formatos = EXCEL_NUMBER_FORMATS_BY_HEADER.get(str(col_name).strip())
        if not formatos:
            continue

        rng = ws_com.Range(ws_com.Cells(first_row, idx), ws_com.Cells(last_row, idx))

        try:
            rng.NumberFormat = formatos[0]
        except Exception:
            pass

        try:
            rng.NumberFormatLocal = formatos[1]
        except Exception:
            pass


def _atualizar_via_com(wb_com, df_out):
    """
    Atualiza dados da aba Base usando Excel COM enquanto o arquivo está aberto.
    Preserva 100%: macros, botões, shapes, personalização, formatação.
    Retorna True se bem-sucedido.
    """
    try:
        ws_com = wb_com.Sheets("Base")

        # Detecta linha de cabeçalho (procura "Nº Nota Fiscal")
        header_row = 2
        for r in range(1, 6):
            for c in range(1, 30):
                try:
                    val = ws_com.Cells(r, c).Value
                    if val == "Nº Nota Fiscal":
                        header_row = r
                        raise StopIteration
                except StopIteration:
                    raise
                except Exception:
                    pass

    except StopIteration:
        pass
    except Exception as e:
        _log(f"[COM] Erro ao acessar aba Base: {e}")
        return False

    sucesso = False
    reprotecao_ok = True
    aba_estava_protegida = False

    try:
        aba_estava_protegida = _desproteger_base_operacional_com(ws_com)
        data_start_row = header_row + 1

        # Calcula área usada real
        used = ws_com.UsedRange
        last_row = used.Row + used.Rows.Count - 1
        last_col = max(used.Column + used.Columns.Count - 1, len(df_out.columns))
        template_last_col = max(last_col, len(df_out.columns))
        template_row_height = None

        # Limpa somente os dados (mantém shapes/formatação)
        # Antes de limpar, preserva o mapa NF → cor da coluna A (status VBA)
        # _norm_nf normaliza 112110 / 112110.0 / "112110" → "112110" (Excel retorna float)
        def _norm_nf(v):
            try:
                return str(int(float(str(v).strip())))
            except Exception:
                return str(v).strip()

        nf_cor_map = {}  # {nf_str: (interior_color, font_color)}
        if last_row >= data_start_row:
            for r in range(data_start_row, last_row + 1):
                try:
                    cell = ws_com.Cells(r, 1)
                    nf_val = cell.Value
                    if nf_val is not None and str(nf_val).strip():
                        nf_str = _norm_nf(nf_val)
                        if nf_str not in nf_cor_map:
                            c_idx = cell.Interior.ColorIndex
                            if c_idx is not None and c_idx != -4142:
                                nf_cor_map[nf_str] = (
                                    cell.Interior.Color,
                                    cell.Font.Color,
                                )
                except Exception:
                    pass

        try:
            template_row_height = ws_com.Rows(data_start_row).RowHeight
        except Exception:
            template_row_height = None

        if last_row >= data_start_row:
            ws_com.Range(
                ws_com.Cells(data_start_row, 1),
                ws_com.Cells(last_row, last_col),
            ).ClearContents()
            # Limpa cores obsoletas da coluna A (fundo E fonte)
            _col_a = ws_com.Range(
                ws_com.Cells(data_start_row, 1),
                ws_com.Cells(last_row, 1),
            )
            _col_a.Interior.ColorIndex = -4142  # xlColorIndexNone
            _col_a.Font.ColorIndex = -4105  # xlColorIndexAutomatic

        # Escreve cabeçalho e detecta coluna da Chave de Acesso
        col_chave = None
        for idx, col_name in enumerate(df_out.columns, start=1):
            ws_com.Cells(header_row, idx).Value = col_name
            if col_name == "Chave de Acesso":
                col_chave = idx

        # Escreve dados em bloco (muito mais rápido que célula a célula)
        rows_data = []
        for _, serie in df_out.iterrows():
            row_vals = []
            for idx, val in enumerate(serie.tolist(), start=1):
                row_vals.append(_valor_excel_limpo(val, como_texto=(idx == col_chave)))
            rows_data.append(row_vals)

        if rows_data:
            n_rows = len(rows_data)
            n_cols = len(rows_data[0])
            format_start_row = max(data_start_row, last_row + 1)

            try:
                if format_start_row <= data_start_row + n_rows - 1:
                    ws_com.Range(
                        ws_com.Cells(data_start_row, 1),
                        ws_com.Cells(data_start_row, template_last_col),
                    ).Copy()
                    ws_com.Range(
                        ws_com.Cells(format_start_row, 1),
                        ws_com.Cells(data_start_row + n_rows - 1, n_cols),
                    ).PasteSpecial(
                        Paste=-4122
                    )  # xlPasteFormats
                    if template_row_height is not None:
                        ws_com.Range(
                            ws_com.Cells(format_start_row, 1),
                            ws_com.Cells(data_start_row + n_rows - 1, 1),
                        ).EntireRow.RowHeight = template_row_height
                    ws_com.Application.CutCopyMode = False
            except Exception:
                pass

            # Formata coluna Chave de Acesso como texto ANTES de escrever
            # (evita que o Excel converta para notação científica)
            if col_chave:
                ws_com.Range(
                    ws_com.Cells(data_start_row, col_chave),
                    ws_com.Cells(data_start_row + n_rows - 1, col_chave),
                ).NumberFormat = "@"

            ws_com.Range(
                ws_com.Cells(data_start_row, 1),
                ws_com.Cells(data_start_row + n_rows - 1, n_cols),
            ).Value = rows_data

            _aplicar_formatos_base_com(
                ws_com,
                df_out.columns,
                data_start_row,
                data_start_row + n_rows - 1,
            )

            # Re-aplica cores de status na coluna A
            # Propaga a cor para TODAS as linhas do grupo (igual ao VBA).
            if nf_cor_map:
                _cur_cor = None
                _cur_font = None
                for row_offset, row_vals in enumerate(rows_data):
                    _is_sep = all(v == "" for v in row_vals)
                    if _is_sep:
                        _cur_cor = None
                        _cur_font = None
                    else:
                        nf_val = row_vals[0]
                        if nf_val is not None and str(nf_val).strip():
                            _key = _norm_nf(nf_val)
                            if _key in nf_cor_map:
                                _cur_cor, _cur_font = nf_cor_map[_key]
                    if _cur_cor is not None and not _is_sep:
                        try:
                            cell = ws_com.Cells(data_start_row + row_offset, 1)
                            cell.Interior.Color = _cur_cor
                            cell.Font.Color = _cur_font
                        except Exception:
                            pass
        wb_com.Save()
        sucesso = True

    except Exception as e:
        _log(f"[COM] Falha ao gravar dados via COM: {e}")
    finally:
        if aba_estava_protegida:
            try:
                _proteger_base_operacional_com(ws_com)
            except Exception as e:
                reprotecao_ok = False
                _log(f"[COM] Falha ao reproteger aba Base: {e}")

    if sucesso and reprotecao_ok:
        _log(
            "[COM] Dados atualizados via Excel COM — "
            "macros, botões e personalização 100% preservados."
        )

    return sucesso and reprotecao_ok


def atualizar_excel_com_espacos(df, ignorar_lancadas=False):
    """Atualiza a aba Base preservando macros, botões e personalização.

    ignorar_lancadas=True: ignora o filtro de nfs_lancadas.txt e traz todos os
    registros do df (usado pelo botão Recarregar para mostrar tudo que está no banco).
    """
    if df.empty:
        return False

    # Remove zeros à esquerda da Série (Origem).
    def normalizar_serie(val):
        try:
            return int(val)
        except (ValueError, TypeError):
            return val

    if "Série (Origem)" in df.columns:
        df["Série (Origem)"] = df["Série (Origem)"].apply(normalizar_serie)

    # Exclui NFs já lançadas e removidas da planilha (evita que voltem após exportação)
    # Quando ignorar_lancadas=True (Recarregar completo), este filtro é ignorado.
    if not ignorar_lancadas:
        _nfs_lancadas = _ler_nfs_lancadas()
        if _nfs_lancadas and "Nº Nota Fiscal" in df.columns:
            df = df[
                ~df["Nº Nota Fiscal"].astype(str).str.strip().isin(_nfs_lancadas)
            ].copy()
            if df.empty:
                _log(
                    "[INFO] Todas as NFs do ciclo atual já foram lançadas. Nada a atualizar."
                )
                return True

    def definir_grupo(row):
        if row["CFOP"] == "5124":
            return row["Nº Nota Fiscal"]
        if row["CFOP"] == "5902" and row["NF Vinculada (5124)"] != "N/A":
            return row["NF Vinculada (5124)"]
        return row["Nº Nota Fiscal"]

    df["Grupo"] = df.apply(definir_grupo, axis=1)
    df = df.sort_values(by=["Grupo", "CFOP", "Seq."], ascending=[False, False, True])

    lista_final = []
    prev_g = None
    for _, row in df.iterrows():
        if prev_g is not None and row["Grupo"] != prev_g:
            lista_final.append({c: None for c in df.columns})
        lista_final.append(row.to_dict())
        prev_g = row["Grupo"]

    df_out = pd.DataFrame(lista_final).drop(columns=["Grupo"])
    df_out = df_out.where(pd.notna(df_out), None)

    # ── Caminho único: Excel COM ───────────────────────────────────────────────
    # Se o arquivo estiver aberto, reutiliza a instância existente.
    # Se estiver fechado, abre a Base em uma instância oculta do Excel.
    # Isso preserva 100%: macros, botões, shapes, imagens, personalização e formatação.
    wb_com_aberto = _wb_aberto_no_excel()
    wb_com = wb_com_aberto
    excel_oculto = None
    abriu_excel_oculto = False

    if wb_com is None:
        excel_oculto, wb_com = _abrir_base_em_excel_invisivel()
        if wb_com is None:
            _log(
                "[ERRO] Não foi possível abrir a Base via Excel COM. Atualização abortada para preservar layout e personalização."
            )
            return False
        abriu_excel_oculto = True
        _log(
            "[COM] Base aberta em instância oculta do Excel para preservar layout e personalização."
        )

    try:
        if _atualizacao_bloqueada_por_status_sap(wb_com):
            _log(
                "[COM] Atualizacao da Base adiada: motor SAP em execucao ou aguardando impressao."
            )
            return None
        if _atualizar_via_com(wb_com, df_out):
            if abriu_excel_oculto:
                _log(
                    "[COM] Base atualizada via Excel oculto — layout, imagens, botões e personalização preservados."
                )
            else:
                _log(
                    "[COM] Base atualizada via Excel já aberto — layout, imagens, botões e personalização preservados."
                )
            return True
        _log(
            "[COM] Atualização via Excel COM falhou. Atualização abortada para preservar layout e personalização."
        )
        return False
    finally:
        if abriu_excel_oculto:
            _fechar_excel_invisivel(excel_oculto, wb_com)


def _carregar_dataframe_base() -> pd.DataFrame:
    if not os.path.exists(caminho_db):
        return pd.DataFrame()

    conn = sqlite3.connect(caminho_db)
    try:
        sujos = conn.execute(
            "SELECT COUNT(*) FROM notas_itens WHERE cfop NOT IN ('5902','5124')"
        ).fetchone()[0]
        if sujos > 0:
            conn.execute("DELETE FROM notas_itens WHERE cfop NOT IN ('5902','5124')")
            conn.commit()
            _log(f"[LIMPEZA] {sujos} registro(s) CFOP invalido removidos.")

        return pd.read_sql_query(
            """
            SELECT
                nf as 'Nº Nota Fiscal',
                seq as 'Seq.',
                cod_material as 'Código Material',
                descricao as 'Descrição do Material',
                ordem_producao as 'Ordem de Produção',
                qtd as 'Qtd.',
                un as 'UN.',
                vlr_unit as 'Vlr. Unitário',
                vlr_total_item as 'Vlr. Total Item',
                cfop as 'CFOP',
                nf_vinculada_5124 as 'NF Vinculada (5124)',
                nf_origem as 'NF Origem do Material',
                serie_origem as 'Série (Origem)',
                vlr_total_nf as 'Vlr. Total (NF)',
                emissao as 'Emissão',
                CASE
                    WHEN lower(trim(COALESCE(chave_acesso, ''))) IN ('nan', 'none', 'nat', '<na>', 'n/a') THEN ''
                    ELSE COALESCE(chave_acesso, '')
                END as 'Chave de Acesso'
            FROM notas_itens
            """,
            conn,
        )
    finally:
        conn.close()


def _sincronizar_base_automatica() -> bool | None:
    global _status_global

    pendentes = _ler_xmls_pendentes_recarregar()
    if pendentes <= 0:
        return True

    df = _carregar_dataframe_base()
    if df.empty:
        _gravar_xmls_pendentes_recarregar(0)
        _status_global = "Banco vazio"
        _log("[PROCESSAMENTO] Banco sem registros para atualizar a Base.")
        return True

    _status_global = f"Atualizando Base com {pendentes} XML(s) novo(s)"
    resultado = atualizar_excel_com_espacos(df)

    if resultado is True:
        _gravar_xmls_pendentes_recarregar(0)
        _status_global = "Base atualizada"
        _log("[PROCESSAMENTO] Base XLSM atualizada automaticamente.")
        return True

    if resultado is None:
        _status_global = "Atualizacao da Base aguardando fim do SAP"
        _log("[PROCESSAMENTO] Atualizacao automatica adiada: SAP em execucao.")
        return None

    _log("[PROCESSAMENTO] Falha na atualizacao automatica da Base.")
    return False


def _atualizar_status_monitor_padrao():
    global _status_global

    status_padrao = _texto_status_monitor()
    if _ler_xmls_pendentes_recarregar() <= 0:
        _status_global = status_padrao
        return

    status_atual = str(_status_global or "").upper()
    if (
        "ATUALIZACAO" in status_atual
        or "FECHE O XLSM" in status_atual
        or "FECHANDO BASE" in status_atual
    ):
        return

    _status_global = status_padrao


# ==========================================
# THREADS DE EXECUÇÃO
# ==========================================


def tarefa_processamento():
    global _status_global, _nf_count_global, _notas_novas_global
    _log("[SISTEMA] Motor de Dados e Relatórios iniciado.")
    inicializar_db()
    arquivos_varridos = {}
    proxima_limpeza_xml = 0.0

    # Carrega chaves já no banco para evitar reprocessar XMLs já conhecidos
    conn = sqlite3.connect(caminho_db)
    chaves_processadas = set(
        row[0] for row in conn.execute("SELECT DISTINCT chave_acesso FROM notas_itens")
    )
    conn.close()
    _nf_count_global = len(chaves_processadas)
    _log(f"[SISTEMA] {_nf_count_global} NF(s) já no banco de dados.")
    _atualizar_status_monitor_padrao()
    _log("[SISTEMA] Atualizacao automatica da Base habilitada.")

    if _ler_xmls_pendentes_recarregar() > 0:
        _sincronizar_base_automatica()

    while True:
        arquivos = [f for f in os.listdir(pasta_xml) if f.lower().endswith(".xml")]
        arquivos_atuais = set(arquivos)
        if arquivos_varridos:
            arquivos_varridos = {
                nome: assinatura
                for nome, assinatura in arquivos_varridos.items()
                if nome in arquivos_atuais
            }
        novos_inseridos = 0

        if arquivos:
            conn = sqlite3.connect(caminho_db)
            for arq in arquivos:
                caminho_arq = os.path.join(pasta_xml, arq)
                try:
                    assinatura_arquivo = (
                        os.path.getsize(caminho_arq),
                        os.path.getmtime(caminho_arq),
                    )
                except OSError:
                    continue

                if arquivos_varridos.get(arq) == assinatura_arquivo:
                    continue

                dados = extrair_detalhes_xml(caminho_arq)
                if dados is None:
                    continue

                if not dados:
                    # XML sem itens CFOP 5902/5124 — ignora silenciosamente
                    _log(f"[IGNORADO] {arq} — sem itens CFOP 5902 ou 5124.")
                    arquivos_varridos[arq] = assinatura_arquivo
                    continue
                chave = dados[0][0]  # chave_acesso é o primeiro campo
                if chave in chaves_processadas:
                    arquivos_varridos[arq] = assinatura_arquivo
                    continue  # já está no banco, ignora
                for item in dados:
                    conn.execute(
                        """
                        INSERT OR REPLACE INTO notas_itens
                            (chave_acesso, nf, seq, cod_material, descricao,
                             ordem_producao, qtd, un, vlr_unit, vlr_total_item,
                             cfop, nf_vinculada_5124, nf_origem, serie_origem,
                             vlr_total_nf, emissao, data_importacao)
                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
                        """,
                        item + (time.strftime("%Y-%m-%d %H:%M:%S"),),
                    )
                chaves_processadas.add(chave)
                arquivos_varridos[arq] = assinatura_arquivo
                novos_inseridos += 1
                _log(f"[PROCESSAMENTO] Nova NF inserida: {arq}")
            conn.commit()

            if novos_inseridos > 0:
                _nf_count_global = len(chaves_processadas)
                _notas_novas_global += novos_inseridos
                pendentes_recarregar = _somar_xmls_pendentes_recarregar(novos_inseridos)
                _status_global = (
                    f"{pendentes_recarregar} XML(s) aguardando atualizacao da Base"
                )
                _log(
                    f"[PROCESSAMENTO] {novos_inseridos} XML(s) novo(s). Iniciando atualizacao automatica da Base."
                )
                _notificar(
                    "RPA Coonagro — XMLs Novos",
                    f"{pendentes_recarregar} XML(s) novo(s). Atualizando a Base automaticamente.",
                )
                conn.close()
            else:
                conn.close()

        if _ler_xmls_pendentes_recarregar() > 0:
            _sincronizar_base_automatica()

        if time.time() >= proxima_limpeza_xml:
            _limpar_xmls_recebidos_antigos(chaves_processadas)
            proxima_limpeza_xml = time.time() + INTERVALO_LIMPEZA_XML_SEGUNDOS

        _atualizar_status_monitor_padrao()

        time.sleep(30)


def _buscar_pasta_outlook(folder, nome, profundidade=0, max_prof=5):
    """Busca recursiva pela pasta com o nome dado."""
    if profundidade > max_prof:
        return None
    try:
        for sub in folder.Folders:
            if sub.Name == nome:
                return sub
            encontrado = _buscar_pasta_outlook(sub, nome, profundidade + 1, max_prof)
            if encontrado:
                return encontrado
    except Exception:
        pass
    return None


def tarefa_outlook():
    _log("[SISTEMA] Coletor Outlook iniciado. Aguardando e-mails...")
    pythoncom.CoInitialize()
    _mostrar_aviso_primeiro_uso_outlook()

    arquivos_salvos = set(
        f.lower() for f in os.listdir(pasta_xml) if f.lower().endswith(".xml")
    )

    while True:
        try:
            outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace(
                "MAPI"
            )
            root_folder = outlook.DefaultStore.GetRootFolder()

            pasta = _buscar_pasta_outlook(root_folder, "XML Coonagro")

            if not pasta:
                _log(
                    "[AVISO] Pasta 'XML Coonagro' não encontrada no Outlook. "
                    "Verifique se a pasta existe e se o Outlook está aberto."
                )
            else:
                itens = pasta.Items
                total = itens.Count
                novos = 0
                for i in range(total, 0, -1):  # reverso para não perder índice
                    try:
                        msg = itens.Item(i)  # Item() usa índice 1-based (padrão COM)
                        # Ignora itens que não são e-mails (reuniões, tarefas, etc.)
                        # MailItem = classe 43
                        if getattr(msg, "Class", None) != 43:
                            continue
                        for anexo in msg.Attachments:
                            nome_arq = anexo.FileName.lower()
                            if (
                                nome_arq.endswith(".xml")
                                and nome_arq not in arquivos_salvos
                            ):
                                destino = os.path.join(pasta_xml, anexo.FileName)
                                anexo.SaveAsFile(destino)
                                arquivos_salvos.add(nome_arq)
                                novos += 1
                                _log(f"[OUTLOOK] XML salvo: {anexo.FileName}")
                        msg.UnRead = False
                    except Exception as e_msg:
                        _log(f"[AVISO] Erro ao processar e-mail #{i}: {e_msg}")

                if novos == 0:
                    _log(f"[OUTLOOK] Verificado ({total} e-mails). Nenhum XML novo.")
                else:
                    _log(f"[OUTLOOK] {novos} XML(s) novo(s) salvo(s).")
                    _notificar(
                        "RPA Coonagro — XMLs Recebidos",
                        (
                            f"{novos} XML(s) baixado(s) do Outlook. "
                            "Validando se entram no fluxo 5902/5124."
                        ),
                    )

        except Exception as e:
            _log(f"[ERRO OUTLOOK] {e}")

        time.sleep(15)


if __name__ == "__main__":
    _log("RPA COONAGRO - VERSÃO PRODUÇÃO iniciada.")

    t1 = threading.Thread(target=tarefa_outlook, daemon=True)
    t2 = threading.Thread(target=tarefa_processamento, daemon=True)
    t1.start()
    t2.start()
    _abrir_painel_monitor()

    if TRAY_ATIVO:

        def _criar_icone():
            icone_base = _carregar_icone_monitor_pillow()
            if icone_base is not None:
                return icone_base

            img = Image.new("RGBA", (64, 64), (0, 0, 0, 0))
            draw = ImageDraw.Draw(img)
            draw.ellipse([2, 2, 62, 62], fill=(26, 120, 56))
            draw.rectangle([18, 14, 46, 50], fill=(255, 255, 255))
            draw.rectangle([22, 20, 42, 24], fill=(26, 120, 56))
            draw.rectangle([22, 29, 42, 33], fill=(26, 120, 56))
            draw.rectangle([22, 38, 42, 42], fill=(26, 120, 56))
            return img

        def _abrir_log(icon, item):
            _abrir_log_monitor()

        def _sair(icon, item):
            _log("Encerrado pelo usuário.")
            icon.stop()

        menu = pystray.Menu(
            pystray.MenuItem(
                lambda item: _texto_pendentes_tray(),
                None,
                enabled=False,
            ),
            pystray.MenuItem(
                lambda item: _texto_extracao_pendente_tray(),
                None,
                enabled=False,
            ),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("Abrir Painel", _abrir_painel_monitor),
            pystray.MenuItem("Abrir XLSM", lambda icon, item: _abrir_base_xlsm()),
            pystray.MenuItem("Ver Log", _abrir_log),
            pystray.Menu.SEPARATOR,
            pystray.MenuItem("Encerrar", _sair),
        )
        icon = pystray.Icon(
            "RPA Coonagro",
            _criar_icone(),
            "RPA Coonagro — XML Monitor",
            menu,
        )
        _icon_ref = icon  # permite que as threads enviem notificações
        icon.run()  # bloqueia thread principal — roda em segundo plano
    else:
        try:
            while True:
                time.sleep(1)
        except KeyboardInterrupt:
            _log("Desligando...")
