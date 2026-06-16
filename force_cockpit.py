"""
force_cockpit.py
Automação do SAP Fiori (Electronic Document Cockpit) via PyAutoGUI + image assets.
Lê as NFs novas do banco de dados gerado pelo XML Monitoring.py e processa
cada uma no cockpit aberto no Edge.

Dependências:
    pip install pyautogui pillow opencv-python

Como usar os assets:
    1. Rode o cockpit manualmente e use capturar_asset("nome_do_elemento")
       para tirar um print recortado de cada botão/campo.
    2. Os arquivos PNG ficam salvos em ./assets/
    3. Use clicar_asset("nome") no fluxo para encontrar e clicar no elemento.
"""

import os
import sys
import io
import time
import sqlite3
import subprocess
import pyautogui
import pyperclip
import win32gui
import win32con
import openpyxl

# Garante que o stdout aceita Unicode mesmo em terminais cp1252 (Windows/VBA Shell).
if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
from openpyxl.styles import PatternFill, Font
from PIL import ImageGrab, Image


# ─────────────────────────────────────────────
# CONFIGURAÇÕES
# ─────────────────────────────────────────────
def _auto_detectar_pasta() -> str:
    base = os.path.join(os.environ.get("USERPROFILE", ""), "The Mosaic Company")
    candidato = os.path.join(base, "Controladoria PGA1 (Arquivos) - RPA - Coonagro")
    if os.path.isdir(candidato):
        return candidato
    if os.path.isdir(base):
        for sub in os.listdir(base):
            if "RPA - Coonagro" in sub and os.path.isdir(os.path.join(base, sub)):
                return os.path.join(base, sub)
    return ""


def _ler_config_cockpit() -> dict:
    import json

    _dir = (
        os.path.dirname(sys.executable)
        if getattr(sys, "frozen", False)
        else os.path.dirname(os.path.abspath(__file__))
    )
    cfg = os.path.join(_dir, "config.json")
    if os.path.exists(cfg):
        try:
            with open(cfg, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            pass
    return {}


def _pasta_runtime_compartilhada() -> str:
    pasta = os.path.join(pasta_trabalho, "system")
    os.makedirs(pasta, exist_ok=True)
    return pasta


def _resolver_arquivo_compartilhado(nome_arquivo: str) -> str:
    caminho_novo = os.path.join(_pasta_runtime_compartilhada(), nome_arquivo)
    caminho_legado = os.path.join(pasta_trabalho, nome_arquivo)

    if os.path.exists(caminho_novo):
        return caminho_novo

    if os.path.exists(caminho_legado):
        try:
            os.replace(caminho_legado, caminho_novo)
            return caminho_novo
        except Exception:
            return caminho_legado

    return caminho_novo


_cfg_c = _ler_config_cockpit()
_cfg_pasta = _cfg_c.get("pasta_trabalho", "")
if "SEU_USUARIO" in _cfg_pasta or not os.path.isdir(_cfg_pasta):
    _cfg_pasta = _auto_detectar_pasta()
pasta_trabalho = (
    _cfg_pasta
    or r"C:\Users\esantan3\The Mosaic Company\Controladoria PGA1 (Arquivos) - RPA - Coonagro"
)
caminho_db = _resolver_arquivo_compartilhado("dados_rpa_coonagro.db")
NOME_XLSX_PREFERIDO = "Base_Operacional_Toll_Coonagro.xlsm"
NOME_XLSX_LEGADO = "Base_Dados_Coonagro.xlsm"


def _resolver_caminho_xlsx() -> str:
    for nome in (NOME_XLSX_PREFERIDO, NOME_XLSX_LEGADO):
        caminho = os.path.join(pasta_trabalho, nome)
        if os.path.exists(caminho):
            return caminho
    return os.path.join(pasta_trabalho, NOME_XLSX_PREFERIDO)


CAMINHO_XLSX = _resolver_caminho_xlsx()
# Arquivo fixo onde o usuário cola os números das NFs a processar (um por linha)
# Se vazio ou inexistente, o script usa o banco de dados como fonte.
NFS_LISTA_FIXA = os.path.join(
    (
        os.path.dirname(sys.executable)
        if getattr(sys, "frozen", False)
        else os.path.dirname(os.path.abspath(__file__))
    ),
    "nfs_para_processar.txt",
)
URL_COCKPIT = (
    "https://mosaic-prod.launchpad.cfapps.us10.hana.ondemand.com/site"
    "?siteId=87162a1d-444e-470f-8041-baaa1f98242d"
    "#ElectronicDocument-cockpit"
    "?sap-ui-app-id-hint=pa1_B60D78B855F8E49DB8283BAE60910677"
    "&sap-ui-tech-hint=GUI"
)

# Pasta onde ficam os screenshots dos elementos da UI
PASTA_ASSETS = os.path.join(
    (
        os.path.dirname(sys.executable)
        if getattr(sys, "frozen", False)
        else os.path.dirname(os.path.abspath(__file__))
    ),
    "assets",
)
os.makedirs(PASTA_ASSETS, exist_ok=True)

# ─ Cores de status na coluna A (ARGB hex para openpyxl) ─────────────────────
# AUTCARR_OK = AZUL | ERRO = VERMELHO | CONCLUIDO = VERDE | SEM_PAR = LARANJA
COR_ERRO = "FFEF5350"  # vermelho
COR_CONCLUIDO = "FF66BB6A"  # verde
COR_AUTCARR = "FF42A5F5"  # azul
COR_SEM_PAR = "FFFF9800"  # laranja

# Segurança: move para canto antes de qualquer clique
pyautogui.FAILSAFE = True
pyautogui.PAUSE = 0.3  # pausa padrão entre ações (segundos)


# ─────────────────────────────────────────────
# BANCO DE DADOS — leitura de NFs
# ─────────────────────────────────────────────
def obter_nfs_pendentes():
    """
    Fonte de NFs a processar:
      1. Se 'nfs_para_processar.txt' tiver conteúdo → usa esse arquivo
         (uma NF por linha; arquivo é limpo após a leitura)
      2. Caso contrário → lê do banco todas as 5124 sem COCKPIT_OK
    """
    # ─ Fonte 1: arquivo fixo ──────────────────────────────────────────────
    if os.path.exists(NFS_LISTA_FIXA):
        with open(NFS_LISTA_FIXA, encoding="utf-8") as f:
            nfs = [linha.strip() for linha in f if linha.strip()]
        if nfs:
            print(
                f"[COCKPIT] {len(nfs)} NF(s) lidas de '{os.path.basename(NFS_LISTA_FIXA)}'."
            )
            # Limpa o arquivo após leitura
            open(NFS_LISTA_FIXA, "w").close()
            return nfs

    # ─ Fonte 2: banco de dados ────────────────────────────────────────────
    conn = sqlite3.connect(caminho_db)
    rows = conn.execute("""
        SELECT DISTINCT nf
        FROM notas_itens
        WHERE cfop = '5124'
          AND COALESCE(status_cockpit, '') != 'COCKPIT_OK'
        ORDER BY emissao
        """).fetchall()
    conn.close()
    return [r[0] for r in rows]


def marcar_cockpit_ok(nf: str):
    """Marca a NF mãe como processada no cockpit."""
    conn = sqlite3.connect(caminho_db)
    conn.execute(
        "UPDATE notas_itens SET status_cockpit = 'COCKPIT_OK' WHERE nf = ? AND cfop = '5124'",
        (nf,),
    )
    conn.commit()
    conn.close()


def garantir_coluna_status_cockpit():
    """Adiciona a coluna status_cockpit ao banco se ainda não existir."""
    conn = sqlite3.connect(caminho_db)
    try:
        conn.execute("ALTER TABLE notas_itens ADD COLUMN status_cockpit TEXT")
        conn.commit()
    except sqlite3.OperationalError:
        pass  # coluna já existe
    conn.close()


def obter_pares_origem(nf_mae: str) -> list:
    """
    Lê o Excel e retorna [nf_baixa, nf_alta] — o menor e o maior valor
    da coluna A (Nº Nota Fiscal) para linhas sem cor na coluna A
    pertencentes ao grupo da NF mãe (col K = nf_mae).
    """
    if not os.path.exists(CAMINHO_XLSX):
        print(f"[COCKPIT] Excel não encontrado: {CAMINHO_XLSX}")
        return []
    try:
        wb = openpyxl.load_workbook(CAMINHO_XLSX, data_only=True, keep_vba=True)
        ws = wb["Base"]
        nfs = []
        for row in ws.iter_rows(min_row=3):
            cel_a = row[0]
            nf_a = str(cel_a.value).strip() if cel_a.value is not None else ""
            nf_vinc = str(row[10].value).strip() if row[10].value is not None else ""
            sem_cor = cel_a.fill is None or cel_a.fill.patternType is None
            if nf_vinc == str(nf_mae).strip() and nf_a and sem_cor:
                if nf_a not in nfs:
                    nfs.append(nf_a)
        wb.close()
        if not nfs:
            return []
        nfs.sort()
        return [nfs[0], nfs[-1]]  # [NF baixa, NF alta]
    except Exception as exc:
        print(f"[COCKPIT] Erro ao ler Excel: {exc}")
        return []


def atualizar_status_excel(nfs_lista: list, status: str, cor_argb: str):
    """
    Pinta a coluna A para cada NF em nfs_lista.
    Silencia PermissionError quando o xlsx está aberto no Excel.
    """
    if not os.path.exists(CAMINHO_XLSX):
        print(f"[STATUS] Excel não encontrado: {CAMINHO_XLSX}")
        return
    try:
        wb = openpyxl.load_workbook(CAMINHO_XLSX, keep_vba=True)
        ws = wb["Base"]
        fill = PatternFill(fill_type="solid", fgColor=cor_argb)
        nfs_set = {str(n).strip() for n in nfs_lista}
        atualizadas = 0
        for row in ws.iter_rows(min_row=3):
            nf_a = str(row[0].value).strip() if row[0].value is not None else ""
            if nf_a in nfs_set:
                row[0].fill = fill
                row[0].font = Font(color="FFFFFFFF", bold=True)
                atualizadas += 1
        wb.save(CAMINHO_XLSX)
        wb.close()
        print(f"[STATUS] {atualizadas} linha(s) coloridas na coluna A -> '{status}'.")
    except PermissionError:
        print(f"[STATUS] Aviso: Excel bloqueado — coluna A não atualizada ({status}).")
    except Exception as exc:
        print(f"[STATUS] Erro ao atualizar coluna A: {exc}")


# ─────────────────────────────────────────────
# NAVEGADOR
# ─────────────────────────────────────────────
def abrir_cockpit_edge():
    """Abre o SAP Fiori Electronic Document Cockpit no Edge e traz para primeiro plano."""
    edge_paths = [
        r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
        r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
    ]
    edge_exe = next((p for p in edge_paths if os.path.exists(p)), None)

    if edge_exe:
        subprocess.Popen([edge_exe, "--start-maximized", URL_COCKPIT])
        print(f"[COCKPIT] Edge aberto via: {edge_exe}")
    else:
        # Fallback: abre pelo comando 'start' do Windows (usa navegador padrão)
        subprocess.Popen(
            ["cmd", "/c", "start", "", "msedge", "--start-maximized", URL_COCKPIT],
            shell=False,
        )
        print("[COCKPIT] Edge aberto via cmd start (fallback).")

    print("[COCKPIT] Aguardando carregamento da página...")
    time.sleep(6)
    _focar_edge()


def _focar_edge():
    """Traz a janela do Edge para o primeiro plano."""
    hwnds = []

    def _enum(hwnd, _):
        if win32gui.IsWindowVisible(hwnd):
            titulo = win32gui.GetWindowText(hwnd)
            if (
                "Edge" in titulo
                or "Cockpit" in titulo
                or "SAP" in titulo
                or "mosaic" in titulo.lower()
            ):
                hwnds.append(hwnd)

    win32gui.EnumWindows(_enum, None)
    if hwnds:
        hwnd = hwnds[0]
        win32gui.ShowWindow(hwnd, win32con.SW_MAXIMIZE)
        try:
            win32gui.SetForegroundWindow(hwnd)
        except Exception:
            pass
        time.sleep(1)
        print("[COCKPIT] Edge em primeiro plano.")
    else:
        print("[COCKPIT] Aviso: janela do Edge não encontrada para focar.")


# ─────────────────────────────────────────────
# IMAGE ASSETS — captura e localização
# ─────────────────────────────────────────────
def _asset(nome: str) -> str:
    """Retorna o caminho completo do asset PNG."""
    if not nome.endswith(".png"):
        nome += ".png"
    return os.path.join(PASTA_ASSETS, nome)


def capturar_asset(nome: str, regiao=None):
    """
    Captura um trecho da tela e salva como asset para uso futuro.

    Uso:
        capturar_asset("btn_pesquisar")
        # ou com região específica (x, y, largura, altura):
        capturar_asset("btn_pesquisar", regiao=(800, 200, 120, 40))

    Depois use clicar_asset("btn_pesquisar") no fluxo.
    """
    if regiao:
        x, y, w, h = regiao
        img = ImageGrab.grab(bbox=(x, y, x + w, y + h))
    else:
        print("Posicione o mouse sobre o elemento e pressione Enter...")
        input()
        mx, my = pyautogui.position()
        # Captura 200×60 px centrada no cursor
        img = ImageGrab.grab(bbox=(mx - 100, my - 30, mx + 100, my + 30))

    caminho = _asset(nome)
    img.save(caminho)
    print(f"[ASSET] Salvo: {caminho}  ({img.size[0]}×{img.size[1]} px)")
    return caminho


def localizar_asset(nome: str, confianca: float = 0.70):
    """
    Procura o asset na tela.
    Retorna pyautogui.Box (left, top, width, height) ou None.
    Usa PIL para abrir o PNG (evita erro de caminho Unicode no OpenCV).
    """
    caminho = _asset(nome)
    if not os.path.exists(caminho):
        print(f"[ASSET] Arquivo não encontrado: {caminho}")
        return None
    try:
        img = Image.open(caminho)
        return pyautogui.locateOnScreen(img, confidence=confianca, grayscale=True)
    except pyautogui.ImageNotFoundException:
        return None


def clicar_asset(
    nome: str, confianca: float = 0.70, timeout: int = 15, botao: str = "left"
) -> bool:
    """
    Aguarda até 'timeout' segundos, localiza o asset na tela e clica no centro.
    botao: 'left' (padrão) ou 'right' para clique direito.
    Retorna True se clicou com sucesso.
    """
    inicio = time.time()
    while time.time() - inicio < timeout:
        loc = localizar_asset(nome, confianca)
        if loc:
            cx, cy = pyautogui.center(loc)
            pyautogui.click(cx, cy, button=botao)
            print(f"[ASSET] Clicou em '{nome}' em ({cx}, {cy}) [botão {botao}]")
            return True
        time.sleep(0.8)
    print(f"[ASSET] '{nome}' não encontrado na tela após {timeout}s.")
    return False


def aguardar_asset(nome: str, confianca: float = 0.70, timeout: int = 15) -> bool:
    """Aguarda até 'timeout' segundos pelo asset aparecer na tela (sem clicar)."""
    inicio = time.time()
    while time.time() - inicio < timeout:
        if localizar_asset(nome, confianca):
            return True
        time.sleep(0.8)
    return False


def digitar(texto: str, intervalo: float = 0.05):
    """Digita texto no campo com foco atual."""
    pyautogui.typewrite(str(texto), interval=intervalo)


def pressionar(tecla: str):
    pyautogui.press(tecla)


# ─────────────────────────────────────────────
# FLUXO DO COCKPIT — todos os pares de uma vez
# ─────────────────────────────────────────────
def processar_todos_no_cockpit() -> bool:
    """
    Executa o fluxo do cockpit UMA VEZ para todos os pares já copiados no clipboard.

    Fluxo (assets em ./assets/):
        1.  option_toll_ind → clica na opção toll/ind
        2.  column_nf       → clique direito na coluna NF
        3.  filter          → abre o painel de filtro
        4.  Tab × 2 + Space → navega e ativa o campo de entrada
        5.  Shift+F12       → cola/abre diálogo de múltiplos valores
        6.  F8              → executa a busca
        7.  select_all      → seleciona todos os resultados
        8.  more_options    → abre menu de opções extras
        9.  ignore_step     → confirma a etapa (Ignorar)
    """
    # ── 1. ──────────────────────────────────────────────────────────────────
    if not clicar_asset("option_toll_ind", timeout=60):
        print("[COCKPIT] Passo 1 falhou: option_toll_ind não encontrado.")
        return False
    time.sleep(2.0)

    # ── 2. ──────────────────────────────────────────────────────────────────
    if not clicar_asset("column_nf", botao="right"):
        print("[COCKPIT] Passo 2 falhou: column_nf não encontrado.")
        return False
    time.sleep(2.0)

    # ── 3. ──────────────────────────────────────────────────────────────────
    if not clicar_asset("filter"):
        print("[COCKPIT] Passo 3 falhou: filter não encontrado.")
        return False
    time.sleep(2.0)

    # ── 4. Tab × 2 + Space ──────────────────────────────────────────────────
    pressionar("tab")
    time.sleep(0.5)
    pressionar("tab")
    time.sleep(0.5)
    pressionar("space")
    time.sleep(1.0)

    # ── 5. Shift+F12 ────────────────────────────────────────────────────────
    pyautogui.hotkey("shift", "f12")
    time.sleep(2.0)

    # ── 6. F8 ───────────────────────────────────────────────────────────────
    pressionar("f8")
    time.sleep(3.0)

    # ── 7. Enter ────────────────────────────────────────────────────────────
    pressionar("enter")
    time.sleep(3.0)

    # ── 8. ──────────────────────────────────────────────────────────────────
    if not clicar_asset("select_all", timeout=30):
        print("[COCKPIT] Passo 8 falhou: select_all não encontrado.")
        return False
    time.sleep(2.0)

    # ── 9. ──────────────────────────────────────────────────────────────────
    if not clicar_asset("more_options"):
        print("[COCKPIT] Passo 9 falhou: more_options não encontrado.")
        return False
    time.sleep(2.0)

    # ── 10. ─────────────────────────────────────────────────────────────────
    if not clicar_asset("ignore_step"):
        print("[COCKPIT] Passo 10 falhou: ignore_step não encontrado.")
        return False
    time.sleep(2.0)

    print("[COCKPIT] Todos os pares processados com sucesso.")
    return True


# ─────────────────────────────────────────────
# MAIN
# ─────────────────────────────────────────────
def main():
    garantir_coluna_status_cockpit()

    nfs = obter_nfs_pendentes()
    if not nfs:
        print("[COCKPIT] Nenhuma NF pendente para processar no cockpit.")
        return

    print(f"[COCKPIT] {len(nfs)} par(es) lidos: {nfs}")

    # ── 0. Coletar TODOS os pares e copiar para clipboard ────────────────
    todos_pares = []
    for nf in nfs:
        if "|" in nf:
            baixa, alta = nf.split("|", 1)
            todos_pares.extend([baixa.strip(), alta.strip()])
        else:
            pares = obter_pares_origem(nf)
            if pares:
                todos_pares.extend(pares)
            else:
                print(
                    f"[COCKPIT] Aviso: sem pares sem cor na coluna A para '{nf}'. Pulando."
                )

    if not todos_pares:
        print("[COCKPIT] Nenhum par de NF encontrado sem cor na coluna A.")
        return

    n_grupos = len(todos_pares) // 2
    print(
        f"[COCKPIT] {n_grupos} grupo(s) -> {len(todos_pares)} NF(s) copiadas para clipboard."
    )
    print(f"[COCKPIT] Pares: {todos_pares}")
    pyperclip.copy("\n".join(todos_pares))

    # ── 0.1 Abrir navegador ──────────────────────────────────────────────
    abrir_cockpit_edge()

    # ── 1-9: Processar TODOS de uma vez no cockpit ───────────────────────
    ok = processar_todos_no_cockpit()

    if ok:
        for nf in nfs:
            nf_db = nf.split("|")[-1].strip() if "|" in nf else nf
            marcar_cockpit_ok(nf_db)
        atualizar_status_excel(todos_pares, "AUTCARR_OK", COR_AUTCARR)
        print(f"[COCKPIT] {len(nfs)} NF(s) marcadas como COCKPIT_OK no banco.")
    else:
        atualizar_status_excel(todos_pares, "ERRO", COR_ERRO)
        print("[COCKPIT] Processamento falhou — NFs não marcadas no banco.")

    print("\n── Resumo ──────────────────────")
    print(
        f"{'Concluido' if ok else 'Erro'}: {n_grupos} grupo(s) | {len(todos_pares)} NF(s)"
    )


if __name__ == "__main__":
    _log_err = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), "cockpit_error.log"
    )
    try:
        main()
    except Exception as _exc:
        import traceback as _tb

        with open(_log_err, "a", encoding="utf-8") as _lf:
            _lf.write(f"\n[{time.strftime('%Y-%m-%d %H:%M:%S')}] ERRO:\n")
            _tb.print_exc(file=_lf)
        raise
